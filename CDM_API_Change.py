import tkinter as tk
from tkinter import *
from tkinter import filedialog
from tkinter.filedialog import askdirectory
import tkinter.ttk as ttk
import pandas as pd
import numpy as np
import traceback
from tkinter import messagebox
import imaplib
import tempfile
import email
import openpyxl
import datetime
import glob
import shutil
import time
from datetime import datetime
import win32com
import win32com.client
import re
import sys
import os
import requests
import json
import pandas as pd
import openpyxl
from geopy.geocoders import Nominatim
from bs4 import BeautifulSoup

# try:
root = Tk()
# root.geometry("1068x768")
root.state('zoomed')
root.title('Provider Search')
# root.resizable(0, 0)
style = ttk.Style()
style.theme_use("default")
style.configure("Treeview", background="Light yellow", foreground="black", fieldbackground="light yellow")
style.map("Treeview", background=[("selected", "green")])
xl = StringVar()
current_var = tk.StringVar()
regex_postfix_removal = '( ,| |,)+(LLC|INC|DBA|PC|PLLC|P.C|LCSW|LMHC|LPC|MFT|MS|LCSW-C|NSS|M.S|M.A|P.L.L.C|PSYDPA)'

def remove_char(ch):
    ch = ch.rstrip(".")
    ch = ch.replace(",", "").replace(".", "").replace("/", "")
    ch = re.sub(regex_postfix_removal, '', ch, re.IGNORECASE)
    return ch

abbr_df = pd.read_excel(os.path.join(os.getcwd(), r'Abbreviation_Sheet.xlsx'))
abbr_df.set_index('abbreviation', inplace=True)

def replace_abbreviations(text):
    if isinstance(text, str):
        words = text.split()
        words = [abbr_df.loc[word, 'expansion'] if word in abbr_df.index else word for word in words]
        text = ' '.join(words)
    return text

def closeBtn():
    root.destroy()

def sheet_data():
    try:
        excel_folder_lst = ["Temp_Files", "Output_File"]
        global df
        df1 = pd.read_excel(input3.get(), sheet_name=current_var.get())
        # print(df1)
        df = pd.DataFrame(df1)


        df = df.apply(lambda x: x.str.upper() if x.dtype == 'object' else x)
        # print(df)
        if 'FIRST NAME' in df.columns:
            df['MIDDLE INITIAL'] = df['MIDDLE INITIAL'].fillna("NO")
            df['MIDDLE INITIAL'] = df['MIDDLE INITIAL'].replace('NO', "")
            df['FIRST NAME'] = df['FIRST NAME'].fillna(" ")
            df[["FIRST NAME", "LAST NAME"]] = df[["FIRST NAME", "LAST NAME"]].astype(str)
            df['PROVIDER'] = df['FIRST NAME'] + ' ' + df['LAST NAME']
            df['FIRST NAME'] = df['FIRST NAME'].str.upper()
            df['LAST NAME'] = df['LAST NAME'].str.upper()
            cols = df.columns.tolist()
            cols = cols[-1:] + cols[:-1]
            df = df[cols]
            df[["NPI", "ZIP", "PHONE"]] = (df[["NPI", "ZIP", "PHONE"]].fillna(0)).astype('int64')
            df[["STREET", "SUITE", "CITY", "STATE"]] = df[
                ["STREET", "SUITE", "CITY", "STATE"]].astype(str)
            dfTemp = pd.DataFrame(df1)
        else:
            df[["NPI", "ZIP", "PHONE"]] = (df[["NPI", "ZIP", "PHONE"]].fillna(0)).astype('int64')
            df[["STREET", "SUITE", "CITY", "STATE"]] = df[
                ["STREET", "SUITE", "CITY", "STATE"]].astype(str)
            dfTemp = pd.DataFrame(df1)
        dfTemp.NPI = dfTemp.NPI.fillna(0).astype(int)
        dfTemp.NPI = dfTemp.NPI.replace(0, '')
        dfTemp = dfTemp.fillna("")
        if len(df) > 0:
            global xyz
            clear_treeview()
            tree["column"] = list(df.columns)
            tree["show"] = "headings"
            for col in tree["column"]:
                tree.heading(col, text=col)
                df_rows = df.to_numpy().tolist()
            for row in df_rows:
                tree.insert("", "end", values=row)
                if xyz == 1:
                    treeScroll = ttk.Scrollbar(tblFrame, orient="vertical", command=tree.yview)
                    treeScrollH = ttk.Scrollbar(tblFrame, orient="horizontal", command=tree.xview)
                    xyz = 2
                else:
                    tree.configure(xscrollcommand=tree.yview)
                    tree.configure(yscrollcommand=tree.yview)
                tree.configure(xscrollcommand=treeScrollH.set, yscrollcommand=treeScroll.set)
                treeScroll.pack(side=LEFT, fill=BOTH)
                treeScrollH.pack(side=BOTTOM, fill=BOTH)
                tree.pack()
                tree.focus_set()
                children = tree.get_children()
                if children:
                    tree.focus(children[0])
                    tree.selection_set(children[0])
    except Exception as e:
        traceback.print_exc(file=sys.stdout)


def clear_treeview():
    tree.delete(*tree.get_children())

def selectItem(e):
    try:
        global x
        global curItem
        curItem = tree.focus()
        x = {}
        x = tree.item(curItem)

    except Exception as e:
        traceback.print_exc(file=sys.stdout)
# def open_file():
#     try:
        # global filepath
        # file = filedialog.askopenfile(mode="r", filetypes=[("Excel Files", "*.xlsx *.csv")])
        # if file:
        #     filepath = os.path.abspath(file.name)
        #     input3.config(state="normal")
        #     input3.insert(0, str(filepath))
        #
        #     wb = openpyxl.load_workbook(filepath)
        #     current_var.set(wb.sheetnames[0])
        #     sheetCombo["values"] = wb.sheetnames
        # else:
        #     input3.config(state="normal")
        #     input3.insert(0, "*** Select a File ***")
        #     input3.config(state="disabled")
        # except Exception as e:
        # traceback.print_exc(file=sys.stdout)


def open_file():
    try:
        # outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        # # Define the email sender and subject to search for
        # account = outlook.Folders("imabhi.in@hotmail.com")
        # email_subject = "Input File"
        #
        # # Iterate through the inbox items
        # mail = account.Folders['Scrapping']  # 6 corresponds to the Inbox folder
        # messages = mail.Items
        #
        # for message in messages:
        #     if message.Subject == email_subject:
        #         # Iterate through attachments
        #         attachments = message.Attachments
        #         Selected_folder = "F:\\Python Code\\POC_API\\CDM code Anisha\\CDM_API\\Download_Attachments"
        #         for attachment in attachments:
        #             if attachment.FileName.endswith('.xlsx'):
        #                 # Save the attachment to a temporary location
        #                 attachment.SaveAsFile(os.path.join(Selected_folder, attachment.FileName))
        global filepath
        file = filedialog.askopenfile(mode="r", filetypes=[("Excel Files", "*.xlsx *.csv")])
        if file:
            filepath = os.path.abspath(file.name)
            input3.config(state="normal")
            input3.insert(0, str(filepath))

            wb = openpyxl.load_workbook(filepath)
            current_var.set(wb.sheetnames[0])
            sheetCombo["values"] = wb.sheetnames
        else:
            input3.config(state="normal")
            input3.insert(0, "*** Select a File ***")
            input3.config(state="disabled")
    except Exception as e:
        traceback.print_exc(file=sys.stdout)
        # Open the Excel file using openpyxl
        # if os.path.exists('attached_file.xlsx'):
        #     input3.config(state="normal")
        #     input3.insert(0, str('attached_file.xlsx'))
        #     wb = openpyxl.load_workbook('attached_file.xlsx')
        #     current_var.set(wb.sheetnames[0])
        #     sheetCombo["values"] = wb.sheetnames
        # else:
        #     input3.config(state="normal")
        #     input3.insert(0, "*** Select a File ***")
        #     input3.config(state="disabled")

    except Exception as e:
        traceback.print_exc(file=sys.stdout)

def exportfile():
    try:
        import ctypes
        folderpath = askdirectory()
        shutil.copy(os.path.join(os.getcwd(), "Output File\Provider_Data_Extracted_Updated_1.xlsx"),
                    folderpath + "//My_Output.xlsx")
        labelx.configure(text="File exported to: " + folderpath + "/My_Output.xlsx")
        ctypes.windll.user32.MessageBoxW(0, "File exported to:" + folderpath, "Information Only", 1)
    except Exception as e:
        traceback.print_exc(file=sys.stdout)


def update_progress_label():
    return f"{pb['value']}%"

def progress():
    if pb['value'] < 100:
        pb['value'] += 20
        root.update_idletasks()
        time.sleep(1)
        value_label['text'] = update_progress_label()


def code_search():
    # try:
        start_time = datetime.now()
        labelz.configure(text="Current Status--Processing", fg="brown")
        root.update()
        output_file = "Output File"
        archive_file = "Archive"
        temp_file = "Temp_Files"
        error_file = "Error Out"
        if getattr(sys, 'frozen', False):
            application_path = os.path.dirname(sys.executable)
            running_mode = "Frozen/executable"
        else:
            try:
                app_full_path = os.path.realpath(__file__)
                application_path = os.path.dirname(app_full_path)
                running_mode = "Non-interative"
            except NameError:
                application_path = os.getcwd()
                running_mode = 'Interactive'
        global output_full_path
        output_full_path = os.path.join(application_path, output_file)
        temp_full_path = os.path.join(application_path, temp_file)
        # driver = webdriver.Chrome(executable_path='' + str(application_path) + '\chromedriver.exe')
        # driver.minimize_window()
        # df['index_col'] = range(1, len(df) + 1)
        # print('\n' + str(filepath) + '_Processing Starts')

    # except Exception as e:
    #     traceback.print_exc(file=sys.stdout)
        # print(df)
        def Nppes():
            try:
                print("*************NPPES Module Has Started**************")
                lst = ["DBA", "Inc", "LLC", "PC", "PLLC", 'Inc.', 'P.C.', 'PLLC']
                lst = [x.upper() for x in lst]

                df['NPI'] = df['NPI'].fillna(0)
                df['NPI'] = df['NPI'].astype(int)

                df['PROVIDER'] = df['PROVIDER'].str.upper()

                if 'FIRST NAME' in df.columns:
                    df_new = df
                    # person_df = df['PROVIDER']
                    org_df = pd.DataFrame()
                else:
                    org_df = df
                    df_new = pd.DataFrame()

                df["NPI"] = df.NPI.map(str)



                combined_data = []  # List to hold combined data
                # print(df)

                # npi_column = df.iloc[:, 3]  # Extract the first columnfirst_column = df.iloc[:, 0]  # Extract the first column
                for i in df_new.index:
                    # print(i)
                    if df.NPI[i] != '0':
                    # driver.get("https://npiregistry.cms.hhs.gov/")  # Open NPPES Website
                    # driver.maximize_window()
                    # entry = org_df.loc[i]
                        city = df['CITY'][i]
                        npi = df['NPI'][i]
                        # print(npi)
                        npi = float(npi)
                        npi = int(npi)
                        npi = str(npi)
                    # npi = npi.replace('0', '')
                    # print(npi)
                    # print(entry)
                        if len(npi) == 10:

                            # for npi_id in npi:
                            #     print(npi)
                                url = f"https://npiregistry.cms.hhs.gov/api/?number={npi}&enumeration_type=&taxonomy_description=&name_purpose=&first_name=&use_first_name_alias=&last_name=&organization_name=&address_purpose=&city=&state=&postal_code=&country_code=&limit=&skip=&pretty=&version=2.1"

                                response = requests.get(url)

                                # Access response content
                                content = response.content

                                # Access response JSON
                                json_data = response.json()

                                # Check response status code
                                status_code = response.status_code

                                # Check response headers
                                headers = response.headers

                                data = json.loads(response.text)

                                providers = data['results']
                                # print(providers)
                                for provider in providers:
                                    npi = provider['number']
                                    try:
                                        display_name = provider['basic']['organization_name']
                                        organizational_subpart = provider['basic']['organizational_subpart']
                                        mailing_address = provider['addresses'][0]['address_1']
                                        mailing_city = provider['addresses'][0]['city']
                                        mailing_state = provider['addresses'][0]['state']
                                        mailing_postal_code = provider['addresses'][0]['postal_code']
                                        mailing_phone_number = provider['addresses'][0]['telephone_number']
                                        primary_address = provider['addresses'][1]['address_1']
                                        primary_city = provider['addresses'][1]['city']
                                        primary_state = provider['addresses'][1]['state']
                                        primary_postal_code = provider['addresses'][1]['postal_code']
                                        if 'telephone_number' in provider['addresses'][1]:
                                            primary_phone_number = provider['addresses'][1]['telephone_number']
                                        else:
                                            primary_phone_number = "N/A"
                                        enumeration_type = provider['enumeration_type']
                                        enumeration_date = provider['basic']['enumeration_date']
                                        last_updated = provider['basic']['last_updated']
                                        status = provider['basic']['status']
                                        authorized_official_first_name = provider['basic']['authorized_official_first_name']
                                        authorized_official_last_name = provider['basic']['authorized_official_last_name']
                                        if 'authorized_official_middle_name' in provider['basic']:
                                            authorized_official_middle_name = provider['basic']['authorized_official_middle_name']
                                        else:
                                            authorized_official_middle_name = "N/A"
                                        authorized_official_telephone_number = provider['basic']['authorized_official_telephone_number']
                                        authorized_official_title_or_position = provider['basic'][
                                            'authorized_official_title_or_position']
                                        specialty = provider['taxonomies'][0]['desc']
                                    except:
                                        f_name = provider['basic']['first_name']
                                        l_name = provider['basic']['last_name']
                                        if 'middle_name' in provider['basic']:
                                            m_name = provider['basic']['middle_name']
                                        else:
                                            m_name = ""
                                        if 'credential' in provider['basic']:
                                            c_name = provider['basic']['credential']
                                        else:
                                            c_name = ""
                                        display_name = f_name + " " + l_name
                                        organizational_subpart = ""
                                        mailing_address = provider['addresses'][0]['address_1']
                                        mailing_city = provider['addresses'][0]['city']
                                        mailing_state = provider['addresses'][0]['state']
                                        mailing_postal_code = provider['addresses'][0]['postal_code']
                                        # mailing_phone_number = provider['addresses'][0]['telephone_number']
                                        if 'telephone_number' in provider['addresses'][0]:
                                            mailing_phone_number = provider['addresses'][0]['telephone_number']
                                        else:
                                            mailing_phone_number = "N/A"
                                        primary_address = provider['addresses'][1]['address_1']
                                        primary_city = provider['addresses'][1]['city']
                                        primary_state = provider['addresses'][1]['state']
                                        primary_postal_code = provider['addresses'][1]['postal_code']
                                        if 'telephone_number' in provider['addresses'][1]:
                                            primary_phone_number = provider['addresses'][1]['telephone_number']
                                        else:
                                            primary_phone_number = "N/A"
                                        enumeration_type = provider['enumeration_type']
                                        enumeration_date = provider['basic']['enumeration_date']
                                        last_updated = provider['basic']['last_updated']
                                        status = provider['basic']['status']
                                        authorized_official_first_name = ""
                                        authorized_official_last_name = ""
                                        if 'authorized_official_middle_name' in provider['basic']:
                                            authorized_official_middle_name = ""
                                        else:
                                            authorized_official_middle_name = ""
                                        authorized_official_telephone_number = ""
                                        authorized_official_title_or_position = ""
                                        specialty = provider['taxonomies'][0]['desc']
                                    # print(specialty)
                                    important_details = {
                                        "NPI": npi, "PROVIDER": display_name, 'ORGANIZATIONAL_SUBPART': organizational_subpart,
                                        "MAILING_ADDRESS": mailing_address, "MAILING_CITY": mailing_city,
                                        "MAILING_STATE": mailing_state, "MAILING_POSTAL_CODE": mailing_postal_code,
                                        "MAILING_PHONE_NUMBER": mailing_phone_number, "PRIMARY_PRACTICE_ADDRESS": primary_address,
                                        "PRIMARY_PRACTICE_CITY": primary_city, "PRIMARY_PRACTICE_STATE": primary_state,
                                        "PRIMARY_PRACTICE_POSTAL_CODE": primary_postal_code,
                                        "PRIMARY_PRACTICE_PHONE_NUMBER": primary_phone_number, "ENUMERATION_TYPE": enumeration_type,
                                        "ENUMERATION_DATE": enumeration_date,
                                        "LAST_UPDATED": last_updated, "STATUS": status,
                                        "AUTHORIZED_OFFICIAL_FIRST_NAME": authorized_official_first_name,
                                        "AUTHORIZED_OFFICIAL_LAST_NAME": authorized_official_last_name,
                                        "AUTHORIZED_OFFICIAL_MIDDLE_NAME": authorized_official_middle_name,
                                        "AUTHORIZED_OFFICIAL_TELEPHONE_NUMBER": authorized_official_telephone_number,
                                        "AUTHORIZED_OFFICIAL_TITLE_OR_POSITION": authorized_official_title_or_position,
                                        "SPECIALTY": specialty}
                                    df_npi = pd.DataFrame.from_dict(important_details, orient='index')
                                    df_npi = df_npi.transpose()
                                    # print(df_npi.to_markdown())
                                    combined_data.append(df_npi)
                    else:
                        entry = df_new.loc[i]
                        city = df['CITY'][i]
                        # print((entry['PROVIDER']))
                        first = entry['FIRST NAME']
                        last = entry['LAST NAME']
                        url = f"https://npiregistry.cms.hhs.gov/api/?number=&enumeration_type=&taxonomy_description=&name_purpose=&first_name={first}&use_first_name_alias=&last_name={last}&organization_name=&address_purpose=&city=&state=&postal_code=&country_code=&limit=&skip=&pretty=&version=2.1"

                        response = requests.get(url)

                        # Access response content
                        content = response.content

                        # Access response JSON
                        json_data = response.json()

                        # Check response status code
                        status_code = response.status_code

                        # Check response headers
                        headers = response.headers

                        data = json.loads(response.text)

                        providers = data['results']
                        # print(providers)
                        for provider in providers:
                            npi = provider['number']
                            try:
                                display_name = provider['basic']['organization_name']
                                # if 'organization_name' in provider['basic']:
                                #     display_name = provider['basic']['organization_name']
                                # else:
                                #     display_name = "N/A"
                                organizational_subpart = provider['basic']['organizational_subpart']
                                # if 'organizational_subpart' in provider['basic']:
                                #     organizational_subpart = provider['basic']['organizational_subpart']
                                # else:
                                #     organizational_subpart = "N/A"
                                mailing_address = provider['addresses'][0]['address_1']
                                mailing_city = provider['addresses'][0]['city']
                                mailing_state = provider['addresses'][0]['state']
                                mailing_postal_code = provider['addresses'][0]['postal_code']
                                mailing_phone_number = provider['addresses'][0]['telephone_number']
                                # if 'telephone_number' in provider['addresses'][0]:
                                #     mailing_phone_number = provider['addresses'][0]['telephone_number']
                                # else:
                                #     mailing_phone_number = "N/A"
                                primary_address = provider['addresses'][1]['address_1']
                                primary_city = provider['addresses'][1]['city']
                                primary_state = provider['addresses'][1]['state']
                                primary_postal_code = provider['addresses'][1]['postal_code']
                                if 'telephone_number' in provider['addresses'][1]:
                                    primary_phone_number = provider['addresses'][1]['telephone_number']
                                else:
                                    primary_phone_number = "N/A"
                                enumeration_type = provider['enumeration_type']
                                enumeration_date = provider['basic']['enumeration_date']
                                last_updated = provider['basic']['last_updated']
                                status = provider['basic']['status']
                                authorized_official_first_name = provider['basic']['authorized_official_first_name']
                                # if 'authorized_official_first_name' in provider['basic']:
                                #     authorized_official_first_name = provider['basic']['authorized_official_first_name']
                                # else:
                                #     authorized_official_first_name = "N/A"
                                authorized_official_last_name = provider['basic']['authorized_official_last_name']
                                # if 'authorized_official_last_name' in provider['basic']:
                                #     authorized_official_first_name = provider['basic']['authorized_official_last_name']
                                # else:
                                #     authorized_official_last_name = "N/A"
                                if 'authorized_official_middle_name' in provider['basic']:
                                    authorized_official_middle_name = provider['basic']['authorized_official_middle_name']
                                else:
                                    authorized_official_middle_name = "N/A"
                                authorized_official_telephone_number = provider['basic'][
                                    'authorized_official_telephone_number']
                                # if 'authorized_official_telephone_number' in provider['basic']:
                                #     authorized_official_telephone_number = provider['basic']['authorized_official_telephone_number']
                                # else:
                                #     authorized_official_telephone_number = "N/A"
                                authorized_official_title_or_position = provider['basic'][
                                    'authorized_official_title_or_position']
                                # if 'authorized_official_title_or_position' in provider['basic']:
                                #     authorized_official_title_or_position = provider['basic']['authorized_official_title_or_position']
                                # else:
                                #     authorized_official_title_or_position = "N/A"
                                specialty = provider['taxonomies'][0]['desc']
                            except:
                                f_name = provider['basic']['first_name']
                                l_name = provider['basic']['last_name']
                                if 'middle_name' in provider['basic']:
                                    m_name = provider['basic']['middle_name']
                                else:
                                    m_name = ""
                                if 'credential' in provider['basic']:
                                    c_name = provider['basic']['credential']
                                else:
                                    c_name = ""
                                display_name = f_name + " " + l_name
                                organizational_subpart = ""
                                mailing_address = provider['addresses'][0]['address_1']
                                mailing_city = provider['addresses'][0]['city']
                                mailing_state = provider['addresses'][0]['state']
                                mailing_postal_code = provider['addresses'][0]['postal_code']
                                # mailing_phone_number = provider['addresses'][0]['telephone_number']
                                if 'telephone_number' in provider['addresses'][0]:
                                    mailing_phone_number = provider['addresses'][0]['telephone_number']
                                else:
                                    mailing_phone_number = "N/A"
                                primary_address = provider['addresses'][1]['address_1']
                                primary_city = provider['addresses'][1]['city']
                                primary_state = provider['addresses'][1]['state']
                                primary_postal_code = provider['addresses'][1]['postal_code']
                                if 'telephone_number' in provider['addresses'][1]:
                                    primary_phone_number = provider['addresses'][1]['telephone_number']
                                else:
                                    primary_phone_number = "N/A"
                                enumeration_type = provider['enumeration_type']
                                enumeration_date = provider['basic']['enumeration_date']
                                last_updated = provider['basic']['last_updated']
                                status = provider['basic']['status']
                                authorized_official_first_name = ""
                                authorized_official_last_name = ""
                                if 'authorized_official_middle_name' in provider['basic']:
                                    authorized_official_middle_name = ""
                                else:
                                    authorized_official_middle_name = ""
                                authorized_official_telephone_number = ""
                                authorized_official_title_or_position = ""
                                specialty = provider['taxonomies'][0]['desc']
                            # print(specialty)
                            important_details = {
                                "NPI": npi, "PROVIDER": display_name, 'ORGANIZATIONAL_SUBPART': organizational_subpart,
                                "MAILING_ADDRESS": mailing_address, "MAILING_CITY": mailing_city,
                                "MAILING_STATE": mailing_state, "MAILING_POSTAL_CODE": mailing_postal_code,
                                "MAILING_PHONE_NUMBER": mailing_phone_number, "PRIMARY_PRACTICE_ADDRESS": primary_address,
                                "PRIMARY_PRACTICE_CITY": primary_city, "PRIMARY_PRACTICE_STATE": primary_state,
                                "PRIMARY_PRACTICE_POSTAL_CODE": primary_postal_code,
                                "PRIMARY_PRACTICE_PHONE_NUMBER": primary_phone_number, "ENUMERATION_TYPE": enumeration_type,
                                "ENUMERATION_DATE": enumeration_date,
                                "LAST_UPDATED": last_updated, "STATUS": status,
                                "AUTHORIZED_OFFICIAL_FIRST_NAME": authorized_official_first_name,
                                "AUTHORIZED_OFFICIAL_LAST_NAME": authorized_official_last_name,
                                "AUTHORIZED_OFFICIAL_MIDDLE_NAME": authorized_official_middle_name,
                                "AUTHORIZED_OFFICIAL_TELEPHONE_NUMBER": authorized_official_telephone_number,
                                "AUTHORIZED_OFFICIAL_TITLE_OR_POSITION": authorized_official_title_or_position,
                                "SPECIALTY": specialty}
                            df_npi = pd.DataFrame.from_dict(important_details, orient='index')
                            df_npi = df_npi.transpose()
                            # print(df_npi.to_markdown())
                            filtered_df = df_npi[
                                (df_npi['PRIMARY_PRACTICE_CITY'] == city)]
                            combined_data.append(filtered_df)
                for i in org_df.index:
                    # print(i)
                    if df.NPI[i] != '0':
                    # driver.get("https://npiregistry.cms.hhs.gov/")  # Open NPPES Website
                    # driver.maximize_window()
                        entry = org_df.loc[i]
                        city = df['CITY'][i]
                        npi = df['NPI'][i]
                        # print(npi)
                        npi = float(npi)
                        npi = int(npi)
                        npi = str(npi)
                        # print(npi)
                        # print(entry)
                        if len(npi) == 10:

                            # for npi_id in npi:
                            #     print(npi)
                                url = f"https://npiregistry.cms.hhs.gov/api/?number={npi}&enumeration_type=&taxonomy_description=&name_purpose=&first_name=&use_first_name_alias=&last_name=&organization_name=&address_purpose=&city=&state=&postal_code=&country_code=&limit=&skip=&pretty=&version=2.1"

                                response = requests.get(url)

                                # Access response content
                                content = response.content

                                # Access response JSON
                                json_data = response.json()

                                # Check response status code
                                status_code = response.status_code

                                # Check response headers
                                headers = response.headers

                                data = json.loads(response.text)

                                providers = data['results']
                                # print(providers)
                                for provider in providers:
                                    npi = provider['number']
                                    try:
                                        display_name = provider['basic']['organization_name']
                                        organizational_subpart = provider['basic']['organizational_subpart']
                                        mailing_address = provider['addresses'][0]['address_1']
                                        mailing_city = provider['addresses'][0]['city']
                                        mailing_state = provider['addresses'][0]['state']
                                        mailing_postal_code = provider['addresses'][0]['postal_code']
                                        mailing_phone_number = provider['addresses'][0]['telephone_number']
                                        primary_address = provider['addresses'][1]['address_1']
                                        primary_city = provider['addresses'][1]['city']
                                        primary_state = provider['addresses'][1]['state']
                                        primary_postal_code = provider['addresses'][1]['postal_code']
                                        if 'telephone_number' in provider['addresses'][1]:
                                            primary_phone_number = provider['addresses'][1]['telephone_number']
                                        else:
                                            primary_phone_number = "N/A"
                                        enumeration_type = provider['enumeration_type']
                                        enumeration_date = provider['basic']['enumeration_date']
                                        last_updated = provider['basic']['last_updated']
                                        status = provider['basic']['status']
                                        authorized_official_first_name = provider['basic']['authorized_official_first_name']
                                        authorized_official_last_name = provider['basic']['authorized_official_last_name']
                                        if 'authorized_official_middle_name' in provider['basic']:
                                            authorized_official_middle_name = provider['basic']['authorized_official_middle_name']
                                        else:
                                            authorized_official_middle_name = "N/A"
                                        authorized_official_telephone_number = provider['basic']['authorized_official_telephone_number']
                                        authorized_official_title_or_position = provider['basic'][
                                            'authorized_official_title_or_position']
                                        specialty = provider['taxonomies'][0]['desc']
                                    except:
                                        f_name = provider['basic']['first_name']
                                        l_name = provider['basic']['last_name']
                                        if 'middle_name' in provider['basic']:
                                            m_name = provider['basic']['middle_name']
                                        else:
                                            m_name = ""
                                        if 'credential' in provider['basic']:
                                            c_name = provider['basic']['credential']
                                        else:
                                            c_name = ""
                                        display_name = f_name + " " + l_name
                                        organizational_subpart = ""
                                        mailing_address = provider['addresses'][0]['address_1']
                                        mailing_city = provider['addresses'][0]['city']
                                        mailing_state = provider['addresses'][0]['state']
                                        mailing_postal_code = provider['addresses'][0]['postal_code']
                                        mailing_phone_number = provider['addresses'][0]['telephone_number']
                                        primary_address = provider['addresses'][1]['address_1']
                                        primary_city = provider['addresses'][1]['city']
                                        primary_state = provider['addresses'][1]['state']
                                        primary_postal_code = provider['addresses'][1]['postal_code']
                                        if 'telephone_number' in provider['addresses'][1]:
                                            primary_phone_number = provider['addresses'][1]['telephone_number']
                                        else:
                                            primary_phone_number = "N/A"
                                        enumeration_type = provider['enumeration_type']
                                        enumeration_date = provider['basic']['enumeration_date']
                                        last_updated = provider['basic']['last_updated']
                                        status = provider['basic']['status']
                                        authorized_official_first_name = ""
                                        authorized_official_last_name = ""
                                        if 'authorized_official_middle_name' in provider['basic']:
                                            authorized_official_middle_name = ""
                                        else:
                                            authorized_official_middle_name = ""
                                        authorized_official_telephone_number = ""
                                        authorized_official_title_or_position = ""
                                        specialty = provider['taxonomies'][0]['desc']
                                    # print(specialty)
                                    important_details = {
                                        "NPI": npi, "PROVIDER": display_name, 'ORGANIZATIONAL_SUBPART': organizational_subpart,
                                        "MAILING_ADDRESS": mailing_address, "MAILING_CITY": mailing_city,
                                        "MAILING_STATE": mailing_state, "MAILING_POSTAL_CODE": mailing_postal_code,
                                        "MAILING_PHONE_NUMBER": mailing_phone_number, "PRIMARY_PRACTICE_ADDRESS": primary_address,
                                        "PRIMARY_PRACTICE_CITY": primary_city, "PRIMARY_PRACTICE_STATE": primary_state,
                                        "PRIMARY_PRACTICE_POSTAL_CODE": primary_postal_code,
                                        "PRIMARY_PRACTICE_PHONE_NUMBER": primary_phone_number, "ENUMERATION_TYPE": enumeration_type,
                                        "ENUMERATION_DATE": enumeration_date,
                                        "LAST_UPDATED": last_updated, "STATUS": status,
                                        "AUTHORIZED_OFFICIAL_FIRST_NAME": authorized_official_first_name,
                                        "AUTHORIZED_OFFICIAL_LAST_NAME": authorized_official_last_name,
                                        "AUTHORIZED_OFFICIAL_MIDDLE_NAME": authorized_official_middle_name,
                                        "AUTHORIZED_OFFICIAL_TELEPHONE_NUMBER": authorized_official_telephone_number,
                                        "AUTHORIZED_OFFICIAL_TITLE_OR_POSITION": authorized_official_title_or_position,
                                        "SPECIALTY": specialty}
                                    df_npi = pd.DataFrame.from_dict(important_details, orient='index')
                                    df_npi = df_npi.transpose()
                                    # print(df_npi.to_markdown())
                                    combined_data.append(df_npi)
                    else:

                        Org_nm_entry = org_df.loc[i, 'PROVIDER']
                        Org_zip_entry = str((org_df.loc[i, 'ZIP']))
                        city = df['CITY'][i]
                        if len(Org_zip_entry) == 4:
                            Org_zip_entry = '0' + str(Org_zip_entry)

                        Org_nm_entry = Org_nm_entry.rstrip(".")
                        Org_nm_entry_final = re.sub(regex_postfix_removal, '', Org_nm_entry, re.IGNORECASE)

                        url = f"https://npiregistry.cms.hhs.gov/api/?number=&enumeration_type=&taxonomy_description=&name_purpose=&first_name=&use_first_name_alias=&last_name=&organization_name={Org_nm_entry_final}&address_purpose=&city=&state=&postal_code=&country_code=&limit=&skip=&pretty=&version=2.1"

                        response = requests.get(url)

                        # Access response content
                        content = response.content

                        # Access response JSON
                        json_data = response.json()

                        # Check response status code
                        status_code = response.status_code

                        # Check response headers
                        headers = response.headers

                        data = json.loads(response.text)

                        providers = data['results']
                        # print(providers)
                        for provider in providers:
                            npi = provider['number']
                            try:
                                display_name = provider['basic']['organization_name']
                                organizational_subpart = provider['basic']['organizational_subpart']
                                mailing_address = provider['addresses'][0]['address_1']
                                mailing_city = provider['addresses'][0]['city']
                                mailing_state = provider['addresses'][0]['state']
                                mailing_postal_code = provider['addresses'][0]['postal_code']
                                mailing_phone_number = provider['addresses'][0]['telephone_number']
                                primary_address = provider['addresses'][1]['address_1']
                                primary_city = provider['addresses'][1]['city']
                                primary_state = provider['addresses'][1]['state']
                                primary_postal_code = provider['addresses'][1]['postal_code']
                                if 'telephone_number' in provider['addresses'][1]:
                                    primary_phone_number = provider['addresses'][1]['telephone_number']
                                else:
                                    primary_phone_number = "N/A"
                                enumeration_type = provider['enumeration_type']
                                enumeration_date = provider['basic']['enumeration_date']
                                last_updated = provider['basic']['last_updated']
                                status = provider['basic']['status']
                                authorized_official_first_name = provider['basic']['authorized_official_first_name']
                                authorized_official_last_name = provider['basic']['authorized_official_last_name']
                                if 'authorized_official_middle_name' in provider['basic']:
                                    authorized_official_middle_name = provider['basic']['authorized_official_middle_name']
                                else:
                                    authorized_official_middle_name = "N/A"
                                authorized_official_telephone_number = provider['basic'][
                                    'authorized_official_telephone_number']
                                authorized_official_title_or_position = provider['basic'][
                                    'authorized_official_title_or_position']
                                specialty = provider['taxonomies'][0]['desc']
                            except:
                                f_name = provider['basic']['first_name']
                                l_name = provider['basic']['last_name']
                                if 'middle_name' in provider['basic']:
                                    m_name = provider['basic']['middle_name']
                                else:
                                    m_name = ""
                                if 'credential' in provider['basic']:
                                    c_name = provider['basic']['credential']
                                else:
                                    c_name = ""
                                display_name = f_name +  " " + l_name
                                organizational_subpart = ""
                                mailing_address = provider['addresses'][0]['address_1']
                                mailing_city = provider['addresses'][0]['city']
                                mailing_state = provider['addresses'][0]['state']
                                mailing_postal_code = provider['addresses'][0]['postal_code']
                                mailing_phone_number = provider['addresses'][0]['telephone_number']
                                primary_address = provider['addresses'][1]['address_1']
                                primary_city = provider['addresses'][1]['city']
                                primary_state = provider['addresses'][1]['state']
                                primary_postal_code = provider['addresses'][1]['postal_code']
                                if 'telephone_number' in provider['addresses'][1]:
                                    primary_phone_number = provider['addresses'][1]['telephone_number']
                                else:
                                    primary_phone_number = "N/A"
                                enumeration_type = provider['enumeration_type']
                                enumeration_date = provider['basic']['enumeration_date']
                                last_updated = provider['basic']['last_updated']
                                status = provider['basic']['status']
                                authorized_official_first_name = ""
                                authorized_official_last_name = ""
                                if 'authorized_official_middle_name' in provider['basic']:
                                    authorized_official_middle_name = ""
                                else:
                                    authorized_official_middle_name = ""
                                authorized_official_telephone_number = ""
                                authorized_official_title_or_position = ""
                                specialty = provider['taxonomies'][0]['desc']
                            # print(specialty)
                            important_details = {
                                "NPI": npi, "PROVIDER": display_name, 'ORGANIZATIONAL_SUBPART': organizational_subpart,
                                "MAILING_ADDRESS": mailing_address, "MAILING_CITY": mailing_city,
                                "MAILING_STATE": mailing_state, "MAILING_POSTAL_CODE": mailing_postal_code,
                                "MAILING_PHONE_NUMBER": mailing_phone_number, "PRIMARY_PRACTICE_ADDRESS": primary_address,
                                "PRIMARY_PRACTICE_CITY": primary_city, "PRIMARY_PRACTICE_STATE": primary_state,
                                "PRIMARY_PRACTICE_POSTAL_CODE": primary_postal_code,
                                "PRIMARY_PRACTICE_PHONE_NUMBER": primary_phone_number, "ENUMERATION_TYPE": enumeration_type,
                                "ENUMERATION_DATE": enumeration_date,
                                "LAST_UPDATED": last_updated, "STATUS": status,
                                "AUTHORIZED_OFFICIAL_FIRST_NAME": authorized_official_first_name,
                                "AUTHORIZED_OFFICIAL_LAST_NAME": authorized_official_last_name,
                                "AUTHORIZED_OFFICIAL_MIDDLE_NAME": authorized_official_middle_name,
                                "AUTHORIZED_OFFICIAL_TELEPHONE_NUMBER": authorized_official_telephone_number,
                                "AUTHORIZED_OFFICIAL_TITLE_OR_POSITION": authorized_official_title_or_position,
                                "SPECIALTY": specialty}
                            df_npi = pd.DataFrame.from_dict(important_details, orient='index')
                            df_npi = df_npi.transpose()
                            filtered_df = df_npi['PRIMARY_PRACTICE_CITY'] == city
                            # print(filtered_df.to_markdown())
                            combined_data.append(filtered_df)

                combined_df = pd.concat(combined_data, axis=0)
                # output_file = temp_full_path + "\_Nppes_doctor_details.xlsx"
                combined_df.to_excel(temp_full_path + "\\Nppes_doctor_details.xlsx", index=False)
                # print(combined_df.to_markdown())
                df_output = combined_df
                # print(df_output['PROVIDER'].to_markdown())


                if len(df_output) == 0:
                    df_output.loc[0] = ["", "", "", "", "", 1]

                # print(df_output.dtypes)
                df_output['PROVIDER'] = df_output['PROVIDER'].str.upper()
                # df['PROVIDER'] = df['PROVIDER'].str.upper()
                df['New'] = df['PROVIDER'].str.strip()
                #
                df['New'] = df['New'].str.upper()

                df_new = df['New']
                df_new = df_new.to_frame(name='New')
                df_new['New'] = df['New'].str.upper()
                # df_new['index_col'] = df['index_col']
                # df_output["Address2"] = df_output["Address2"].fillna("Nan")
                # df_output["Phone2"] = df_output["Phone2"].fillna("Nan")
                # df_output["ADDRESS"] = df_output["ADDRESS"] + "'],[' " + df_output["Address2"]
                # df_output["Phone"] = df_output["Phone"] + " '], [' " + df_output["Phone2"]
                final = pd.concat([df, df_new], axis=1)
                final = final.loc[:, ~final.columns.duplicated()]
                df_output["New"] = df_output['PROVIDER']

                df_output["New"] = df_output['New'].str.upper()
                df['New'] = df.New.apply(lambda ch: remove_char(ch))


                if ('FIRST NAME' in df.columns):
                    Vlookup = pd.merge(df_new, df_output, on='New', how='inner')  # Inner Join
                    final_output = Vlookup.drop_duplicates()
                    df_final = pd.merge(final, final_output, on='New', how='left')

                    df_final = df_final.drop(["New"], axis=1)
                    df_final.rename(columns={'STATE_x': 'STATE', 'PROVIDER_x': 'PROVIDER'}, inplace=True)
                    # print(final_output.to_markdown())
                    #
                    # df_output["New"] = df_output.New.apply(lambda ch: remove_char(ch))
                    #
                    # pat_lname = f"({'|'.join(df['LAST NAME'])})"
                    # pat_fname = f"({'|'.join(df['FIRST NAME'])})"
                    # df_output['LAST NAME'] = df_output['New'].str.extract(pat=pat_lname)
                    # df_output['FIRST NAME'] = df_output['New'].str.extract(pat=pat_fname)
                    # df_output.loc[df_output['LAST NAME'].isna(), 'LAST NAME'] = np.NAN
                    # df_output.loc[df_output['FIRST NAME'].isna(), 'FIRST NAME'] = np.NAN
                    # Vlookup = df.merge(df_output, left_on=['LAST NAME', 'FIRST NAME'],
                    #                    right_on=['LAST NAME', 'FIRST NAME'], how='left')
                    # Vlookup = Vlookup.drop(["PROVIDER_y", "New_x", "New_y"], axis=1)
                    #
                    # df_final = Vlookup.drop_duplicates()
                    # df_final.rename(columns={'PROVIDER_x': 'PROVIDER'}, inplace=True)
                    #
                    df_ph = df_final.groupby('PROVIDER')["PRIMARY_PRACTICE_PHONE_NUMBER"].apply(lambda x: x.unique()).reset_index()

                    df_add = df_final.groupby('PROVIDER')["MAILING_ADDRESS"].apply(lambda x: x.unique()).reset_index()

                    df_Nppes = pd.merge(df, df_ph, on='PROVIDER', how='left')

                    df_Nppes = pd.merge(df_Nppes, df_add, on='PROVIDER', how='left')
                    df_Nppes = df_Nppes.drop(['New'], axis=1)
                    df_Nppes.rename(columns={'PRIMARY_PRACTICE_PHONE_NUMBER': 'NPPES_PHONE_NO'}, inplace=True)

                    df_Nppes.rename(columns={'MAILING_ADDRESS': 'NPPES_ADDRESS'}, inplace=True)
                    df_Nppes = df_Nppes.fillna("")
                    df_Nppes.to_excel(temp_full_path + '\\Matched_NPPES_Output.xlsx', index=False, header=True)
                    # print(df_Nppes.to_markdown())
                else:
                    # df_new["New"] = df_new["New"].replace('\s+', '', regex=True)
                    # df_output["New"] = df_output["New"].replace('\s+', '', regex=True)
                    # df['New'] = df['PROVIDER'].replace('\s+', '', regex=True)

                    # df_new['New'] = df_new.New.apply(lambda ch: remove_char(ch))
                    # df_output["New"] = df_output.New.apply(lambda ch: remove_char(ch))
                    # df['New'] = df.New.apply(lambda ch: remove_char(ch))

                    # final = pd.concat([df, df_new], axis=1)
                    # final = final.loc[:, ~final.columns.duplicated()]
                    # print(df_new.to_markdown())
                    # print(df_output.to_markdown())
                    Vlookup = pd.merge(df_new, df_output, on=['New'], how='inner')
                    final_output = Vlookup.drop_duplicates()

                    df_final = pd.merge(final, final_output, on=['New'], how='left')

                    df_final = df_final.drop(["New"], axis=1)
                    df_final.rename(columns={'STATE_x': 'STATE', 'PROVIDER_x': 'PROVIDER'}, inplace=True)

                    df_ph = df_final.groupby("PROVIDER")["PRIMARY_PRACTICE_PHONE_NUMBER"].apply(lambda x: x.unique()).reset_index()

                    df_add = df_final.groupby("PROVIDER")["PRIMARY_PRACTICE_ADDRESS"].apply(lambda x: x.unique()).reset_index()

                    df_Nppes = pd.merge(df, df_ph, on='PROVIDER', how='left')

                    df_Nppes = pd.merge(df_Nppes, df_add, on='PROVIDER', how='left')
                    df_Nppes = df_Nppes.drop(['New'], axis=1)
                    df_Nppes.rename(columns={'PRIMARY_PRACTICE_PHONE_NUMBER': 'NPPES_PHONE_NO'}, inplace=True)

                    df_Nppes.rename(columns={'PRIMARY_PRACTICE_ADDRESS': 'NPPES_ADDRESS'}, inplace=True)
                    df_Nppes = df_Nppes.fillna("")
                    # print(df_Nppes.to_markdown())
                    df_Nppes.to_excel(temp_full_path + '\\Matched_NPPES_Output.xlsx', index=False, header=True)
                    # print(df_Nppes.to_markdown())
                end_time = datetime.now()
                print("NPPES Website Module completed successfully!")
                # Save the DataFrame to an Excel file
                labela.configure(text="NPPES Website Module completed!")
                progress()
                root.update()
            except Exception as e:
                labela.configure(text="NPPES Website Module completed!", fg="brown")
                progress()
                root.update()
                traceback.print_exc(file=sys.stdout)



        def Hippaspace():

            try:
                print("\n****Hippa Space Module has started****")
                # print(df.to_markdown())
                # global df
                # df1 = pd.read_excel(input3.get(), sheet_name=current_var.get())
                # # print(df1)
                # df1["PROVIDER"] = df1['FIRST NAME'] + " " + df1['LAST NAME']
                # df = pd.DataFrame(df1)
                # print(df.to_markdown())
                # # Lst = ["DBA", "INC", "LLC", "PLLC", "PC", "SC", "Inc", "P.C."]
                # if 'FIRST NAME' in df.columns:
                #     individual_df = df['PROVIDER']
                # else:
                #     org_df = df['PROVIDER']
                #     individual_df = pd.DataFrame()

                df_nppes = pd.read_excel("F:\\Python Code\\POC_API\\CDM code Anisha\\CDM_API\\Temp_Files\\Nppes_doctor_details.xlsx") #change path
                # print(df1)
                df_hippa = pd.DataFrame(df_nppes)
                token = "F79E6496440449A88A518210A75FF8EECD2F9297F1EF4EB589CFA8CC759275F4"
                #"A5B45E20BD3A4E7FAD14E1D7BDC015F9E0EC6139858940B789822748A9ADECDF"



                # Read NPI IDs from Excel file
                # df_input = pd.read_excel(input_file)
                # npi_ids = df_input.iloc[:, 0].tolist()  # Assumes NPI IDs are in the first column

                combined_data = []  # List to hold combined data

                for npi_id in df_hippa['NPI']:
                    endpoint = f"https://www.hipaaspace.com/api/npi/getcode?q={npi_id}&token={token}"
                    response = requests.get(endpoint)
                    response.raise_for_status()

                    data = json.loads(response.text)
                    # print(data)
                    df_h = pd.DataFrame(data['NPI'])
                    combined_data.append(df_h)

                combined_df = pd.concat(combined_data, axis=0)
                # print(combined_df.to_markdown())
                if ('FIRST NAME' in df.columns):
                    combined_df['PROVIDER'] = combined_df['FirstName'] + " " + combined_df['LastName']
                else:
                    combined_df['PROVIDER'] = combined_df['OrgName']


                output_file = "F:\\Python Code\\POC_API\\Hipaaspace\\Hipaaspace_doctor_details.xlsx"
                combined_df.to_excel(output_file, index=False)
                # print(combined_df.to_markdown())
                # combined_df['PROVIDER'] = combined_df['FirstName'] + " " + combined_df['LastName']
                df_output = combined_df
                df_output['PROVIDER'] = df_output['PROVIDER'].str.upper()
                # df['PROVIDER'] = df['PROVIDER'].apply(replace_abbreviations)
                # List1 = ["DBA", "INC", "LLC", "PLLC", "PC", "SC", "LMHC", ".", "P.C", "Inc", ","]
                df['PROVIDER'] = df['PROVIDER'].str.upper()
                df['PROVIDER'] = df['PROVIDER'].str.upper()
                df['PROVIDER'] = df['PROVIDER'].apply(replace_abbreviations)
                df['New'] = df['PROVIDER'].str.strip()
                df['New'] = df['New'].str.upper()
                # df['New'] = df['PROVIDER'].replace('\s+', '', regex=True)
                # df['New'] = df.New.apply(lambda ch: remove_char(ch))
                df['New'] = df['New'].str.upper()

                df_new = df['New']
                df_new = df_new.to_frame(name='New')
                # df_new['New'] = df['New'].str.upper()
                # df_new['index_col'] = df['index_col']
                # df['New'] = df['NEW'].replace('\s+', '', regex=True)
                # df['New'] = df.New.apply(lambda ch: remove_char(ch))

                final = pd.concat([df, df_new], axis=1)
                #
                final = final.loc[:, ~final.columns.duplicated()]
                # print(final.to_markdown())
                df_output["New"] = df_output['PROVIDER']
                df_output["New"] = df_output["New"].astype(str)
                # print(df_output.to_markdown())
                if ('FIRST NAME' in df.columns):
                    Vlookup = pd.merge(df_new, df_output, on='New', how='inner')  # Inner Join
                    final_output = Vlookup.drop_duplicates()

                    df_final = pd.merge(final, final_output, on=['New'], how="left")

                    df_final = df_final.drop(["New"], axis=1)

                    df_final.rename(columns={'STATE_x': 'STATE', 'PROVIDER_x': 'PROVIDER'}, inplace=True)

                    df_ph = df_final.groupby("PROVIDER")["PracticeLocationAddressTelephoneNumber"].apply(lambda x: x.unique()).reset_index()

                    df_add = df_final.groupby("PROVIDER")["FirstLinePracticeLocationAddress"].apply(lambda x: x.unique()).reset_index()

                    df_hippa1 = pd.merge(df, df_ph, on='PROVIDER', how='left')

                    df_hippa1 = pd.merge(df_hippa1, df_add, on='PROVIDER', how='left')
                    df_hippa1 = df_hippa1.drop(['New'], axis=1)
                    df_hippa1.rename(columns={'PracticeLocationAddressTelephoneNumber': 'HIPPASPACE_PHONE_NO'}, inplace=True)

                    df_hippa1.rename(columns={'FirstLinePracticeLocationAddress': 'HIPPASPACE_ADDRESS'}, inplace=True)
                    df_hippa1 = df_hippa1.fillna("")

                    df_hippa1.to_excel(temp_full_path + '\Matched_Hippa_Org_Output.xlsx', index=False, header=True)
                    # print(df_hippa1.to_markdown())
                else:
                    # df_new["New"] = df_new["New"].replace('\s+', '', regex=True)
                    # df_output["New"] = df_output["New"].replace('\s+', '', regex=True)
                    # df['New'] = df['PROVIDER'].replace('\s+', '', regex=True)
                    #
                    # df_new['New'] = df_new.New.apply(lambda ch: remove_char(ch))
                    # df_output["New"] = df_output.New.apply(lambda ch: remove_char(ch))
                    # df['New'] = df.New.apply(lambda ch: remove_char(ch))
                    #
                    # final = pd.concat([df, df_new], axis=1)
                    # final = final.loc[:, ~final.columns.duplicated()]

                    Vlookup = pd.merge(df_new, df_output, on='New', how='inner')  # Inner Join
                    final_output = Vlookup.drop_duplicates()
                    # print(final_output.to_markdown())
                    df_final = pd.merge(final, final_output, on=['New'], how="left")
                    # print(df_final.to_markdown())
                    df_final = df_final.drop(["New"], axis=1)

                    df_final.rename(columns={'STATE_x': 'STATE', 'PROVIDER_x': 'PROVIDER'}, inplace=True)

                    df_ph = df_final.groupby("PROVIDER")["PracticeLocationAddressTelephoneNumber"].apply(
                        lambda x: x.unique()).reset_index()

                    df_add = df_final.groupby("PROVIDER")["FirstLinePracticeLocationAddress"].apply(
                        lambda x: x.unique()).reset_index()

                    df_hippa1 = pd.merge(df, df_ph, on='PROVIDER', how='left')

                    df_hippa1 = pd.merge(df_hippa1, df_add, on='PROVIDER', how='left')
                    df_hippa1 = df_hippa1.drop(['New'], axis=1)
                    df_hippa1.rename(columns={'PracticeLocationAddressTelephoneNumber': 'HIPPASPACE_PHONE_NO'},
                                     inplace=True)

                    df_hippa1.rename(columns={'FirstLinePracticeLocationAddress': 'HIPPASPACE_ADDRESS'}, inplace=True)
                    df_hippa1 = df_hippa1.fillna("")

                    df_hippa1.to_excel(temp_full_path + '\Matched_Hippa_Org_Output.xlsx', index=False, header=True)
                    # print(df_hippa1.to_markdown())
                end_time = datetime.now()

                print("Hippa Space Website Module completed successfully!")

                labelb.configure(text="Hippa Space Website Module completed!")
                progress()
                root.update()

            except Exception as e:
                labelb.configure(text="Hippa Space Website Module completed!", fg="brown")
                progress()
                root.update()
                traceback.print_exc(file=sys.stdout)

        def Webmd():
            try:

                print("\n****Webmd Module has started****")

                def get_coordinates(zipcode):
                    geolocator = Nominatim(user_agent="zipcode_converter")
                    location = geolocator.geocode(zipcode)

                    if location:
                        latitude = location.latitude
                        longitude = location.longitude
                        return latitude, longitude
                    else:
                        return None

                df1 = pd.read_excel(input3.get(), sheet_name=current_var.get())
                # # print(df1)
                if 'FIRST NAME' in df1.columns:
                    df1["PROVIDER"] = df1['FIRST NAME'] + " " + df1['LAST NAME']
                    df = pd.DataFrame(df1)
                else:
                    df = pd.DataFrame(df1)
                # Output DataFrame
                df_output = pd.DataFrame()
                # zipcode = "32837"  # Replace with the desired ZIP code
                # coordinates = get_coordinates(zipcode)
                for index, row in df.iterrows():
                    provider_name = row["PROVIDER"]
                    zipcode = str(row["ZIP"])
                    # print(zipcode)
                    # print(provider_name)
                    coordinates = get_coordinates(zipcode)
                    if coordinates:
                        latitude, longitude = coordinates
                        latitude = str(latitude)
                        longitude = str(longitude)
                        # print(latitude)
                        # print(longitude)
                        lat = latitude
                        lon = longitude
                        cor = lat + ", " + lon
                    else:
                        print("Coordinates not found.")

                    url = "https://www.webmd.com/search/2/api/lhd_v_search"
                    Name = provider_name
                    Name_split = Name.split(",")
                    # print(Name_split)
                    if "Dr." in Name_split[0]:
                        Name_split = Name_split[0].replace("Dr. ", "")
                        Name_split = Name_split.split(' ')
                        first_name = Name_split[0]
                        last_name = Name_split[-1]
                        # print(first_name)
                        # print(last_name)
                    else:
                        Name = Name.rstrip()
                        # print(Name)

                    querystring = {"sortby": "bestmatch", "entity": "all", "distance": "40", "newpatient": "", "isvirtualvisit": "",
                                   "minrating": "0", "start": "0", "q": Name, "pt": cor, "specialtyid": "", "d": "40", "sid": "", "pid": "",
                                   "insuranceid": "", "exp_min": "min", "exp_max": "max"}
                    headers = {
                        "cookie": "__cfruid=a389e0f5d53668cc8628f9013d32f2f22c13946a-1686919328",
                        "authority": "www.webmd.com",
                        "accept": "application/json, text/plain, */*",
                        "accept-language": "en-US,en;q=0.9",
                        "client_id": "e4e3f73a-0ceb-4d37-939e-90ddb1238360",
                        "enc_data": "Vt6xzQLzXz4/D/u9afOIhrTrVrKB9wdWL1xhRPwher4=",
                        "if-modified-since": "Fri, 16 Jun 2023 12:36:49 GMT",
                        "origin": "https://doctor.webmd.com",
                        "referer": "https://doctor.webmd.com/",
                        "sec-ch-ua-mobile": "?0",
                        "sec-fetch-dest": "empty",
                        "sec-fetch-mode": "cors",
                        "sec-fetch-site": "same-site",
                        "timestamp": "Fri, 16 Jun 2023 12:37:35 GMT",
                        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36 Edg/114.0.1823.43",
                        "Content-Type": "application/json"
                    }
                    response = requests.request("GET", url, headers=headers, params=querystring)
                    json_data = json.loads(response.text)
                    # print((json_data['data']['response']))
                    data = json_data['data']['response']
                    df_web = pd.DataFrame(data)
                    # df_web['PROVIDER'] = df_web['PROVIDER'].str.upper()
                    # df_web['PROVIDER'] = df_web['firstname'] + " " + df_web['lastname']
                    # df_web['PROVIDER'] = df_web['PROVIDER'].str.upper()
                    # print(df_web.to_markdown())

                    # df_web.to_excel("F:\\Python Code\\POC_API\\CDM code Anisha\\Temp_Files\\webmd.xlsx")
                    filtered_df_non = {'providerid': "", 'firstname': "", 'middlename': "", 'lastname': "", 'degreeabbr': "",
                                       'displayspecialty_mvs': "", 'suffix': "", 'npi': "",
                                       'photourl': "", 'bio_s': "", 'grad_year_d': "", 'q_yearofgraduation': "",
                                       'review_count_d': "", 'c1_avg_f': "", 'c2_avg_f': "", 'c3_avg_f': "",
                                       'c4_avg_f': "", 'specialty_consumername_mvs': "", 'id': "", 'appointmentlinkurl_s': "",
                                       'profiletype_s': "", 'providerlinkurl_s': "",
                                       'provider_url_s': "", 'provider_enhancedbookappointmenturl_s': "", 'location_nimvs': "",
                                       'review_nimvs': "",
                                       'campaignoverrrides_nis': "", 'provider_awards_nis': "", 'dynamic_supporttelehealth_b': "",
                                       'pagedata_nis': "", 'fullname': "",
                                       'providerurl': "", 'awards_count': "", 'Newpatient': "", 'graduationyYear': "",
                                       'yearsofexperience': "", "PHONE_NUMBER": "", "ADDRESS":"", "PROVIDER": ""}

                    if df_web.empty:
                        df_output = df_output._append(filtered_df_non, ignore_index=True)
                    else:
                        if df_web['firstname'].isnull().empty:
                            df_web['PROVIDER'] = ""
                        else:
                            df_web['PROVIDER'] = df_web['firstname'] + " " + df_web['lastname']
                        # print(df_web['PROVIDER'].to_markdown())
                        # print(Name)
                        filtered_df = df_web[(df_web['PROVIDER'] == Name)]
                        # print(filtered_df.to_markdown())
                        # for n in df_output['location_nimvs'][0]:
                        #     print(n)
                        try:
                            location_nimvs = (filtered_df['location_nimvs'][0])
                            # print(location_nimvs[0])
                            data_dict = json.loads(location_nimvs[0])

                            # print(data_dict['LocationPhone'])
                            # print(data_dict['address'])
                            filtered_df["PHONE_NUMBER"] = data_dict['LocationPhone']
                            filtered_df["ADDRESS"] = data_dict['address']
                        except:
                            filtered_df["PHONE_NUMBER"] = ""
                            filtered_df["ADDRESS"] = ""
                        if filtered_df.empty:
                            df_output = df_output._append(filtered_df_non, ignore_index=True)
                        else:
                            # data = {}
                            df_output = df_output._append(filtered_df, ignore_index=True)
                # output_file.drop_duplicates(inplace=True)
                output_file = "F:\\Python Code\\POC_API\\Webmd\\Webmd_doctor_details.xlsx"
                df_output.to_excel(output_file, index=False)
                # print(df_output.to_markdown())

                df['PROVIDER'] = df['PROVIDER'].str.upper()
                # df['PROVIDER'] = df['PROVIDER'].apply(replace_abbreviations)
                df['New'] = df['PROVIDER'].str.strip()
                # df['New'] = df['New'].str.upper()
                # df['New'] = df['PROVIDER'].replace('\s+', '', regex=True)
                # df['New'] = df.New.apply(lambda ch: remove_char(ch))

                df_new = df['New']
                df_new = df_new.to_frame(name='New')
                df_new['New'] = df['New'].str.upper()
                # df_new['index_col'] = df['index_col']
                # df_new["New"] = df_new["New"].replace('\s+', '', regex=True)
                # df_new['New'] = df_new.New.apply(lambda ch: remove_char(ch))

                final = pd.concat([df, df_new], axis=1)

                final = final.loc[:, ~final.columns.duplicated()]
                # print(final.to_markdown())
                # print(df_output.to_markdown())
                df_output['New'] = df_output['PROVIDER']
                df_output['New'] = df_output['New'].astype(str)
                df_output['New'] = df_output['New'].str.upper()
                # df_output["New"] = df_output["New"].replace('\s+', '', regex=True)
                # df_output["New"] = df_output.New.apply(lambda ch: remove_char(ch))
                # print(df_new)
                # print(df_output.to_markdown())
                Vlookup = pd.merge(df_new, df_output, on=['New'], how='inner')
                final_output = Vlookup

                df_final = pd.merge(final, final_output, on=['New'], how='left')

                df_final = df_final.drop(["New"], axis=1)
                df_final.rename(columns={'STATE_x': 'STATE', 'PROVIDER_x': 'PROVIDER'}, inplace=True)
                # print(df_final.to_markdown())
                df_ph = df_final.groupby('PROVIDER')["PHONE_NUMBER"].apply(lambda x: x.unique()).reset_index()

                df_add = df_final.groupby('PROVIDER')["ADDRESS"].apply(lambda x: x.unique()).reset_index()

                df_webmd1 = pd.merge(df, df_ph, on='PROVIDER', how='left')

                df_webmd1 = pd.merge(df_webmd1, df_add, on='PROVIDER', how='left')
                df_webmd1 = df_webmd1.drop(['New'], axis=1)
                df_webmd1.rename(columns={'PHONE_NUMBER': 'WEBMD_PHONE_NO'}, inplace=True)

                df_webmd1.rename(columns={'ADDRESS': 'WEBMD_ADDRESS'}, inplace=True)
                # print(df_webmd1)

                df_webmd1.to_excel(temp_full_path + '\Matched_Webmd_Org_Output.xlsx', index=False, header=True)
    # filtered_df.to_excel('F:\\Python Code\\POC_API\\Webmd\\Webmd_doctor_details.xlsx', index=False)
                print("Webmd Website Module completed!")
                labelc.configure(text="Webmd Website Module completed!")
                progress()
                root.update()
    #
#
#
            except Exception as e:
                labelc.configure(text="Webmd Website Module completed!", fg="brown")
                progress()
                root.update()
                traceback.print_exc(file=sys.stdout)


        def Healthgrades():
            try:
                print("\n****Healthgrades Module has started***")
                df_output = []
                def get_coordinates(zipcode):
                    geolocator = Nominatim(user_agent="zipcode_converter")
                    location = geolocator.geocode(zipcode)

                    if location:
                        latitude = location.latitude
                        longitude = location.longitude
                        return latitude, longitude
                    else:
                        return None


                # Read input file


                df1 = pd.read_excel(input3.get(), sheet_name=current_var.get())
                # # print(df1)
                if 'FIRST NAME' in df1.columns:
                    df1["PROVIDER"] = df1['FIRST NAME'] + " " + df1['LAST NAME']
                    df = pd.DataFrame(df1)
                else:
                    df = pd.DataFrame(df1)
                df_output = pd.DataFrame()

                for index, row in df.iterrows():
                    # try:
                    provider_name = row["PROVIDER"].rstrip()
                    zipcode = str(row["ZIP"])

                    coordinates = get_coordinates(zipcode)

                    if coordinates:
                        latitude, longitude = coordinates
                        latitude = str(latitude)
                        longitude = str(longitude)
                        lat = latitude
                        lon = longitude
                        cor = lat + ", " + lon
                    else:
                        print("Coordinates not found.")

                    url = "https://www.healthgrades.com/api3/usearch"

                    # session_id = str(uuid.uuid4())
                    # request_id = str(uuid.uuid4())

                    querystring = {
                        "where": zipcode,
                        "pt": cor,
                        "sort.provider": "bestmatch",
                        "what": provider_name,
                        "highlight": provider_name,
                        "category": "provider",
                        "cid": "",
                        "hgTrace": "false",
                        "isPsr": "false",
                        "isFsr": "false",
                        "isFirstRequest": "true",
                        "pageNum": "1",
                        "userLocalTime": "15:09"
                    }

                    payload = ""
                    headers = {
                        "cookie": "<your_cookie_value>",
                        "authority": "www.healthgrades.com",
                        "accept": "*/*",
                        "accept-language": "en-US,en;q=0.9,hi;q=0.8",
                        "sec-ch-ua-mobile": "?0",
                        "sec-fetch-dest": "empty",
                        "sec-fetch-mode": "cors",
                        "sec-fetch-site": "same-origin",
                        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"
                    }

                    response = requests.request("GET", url, data=payload, headers=headers, params=querystring)
                    data = json.loads(response.text)
                    providers = data['search']['searchResults']['provider']['results']
                    # print(providers)
                    for provider in providers:
                        # print(provider)
                        # name = provider['displayName']
                        # name = name[0]
                        # print(provider_name)
                        # if name == provider_name:
                        display_name = provider['displayName']

                        display_name = display_name.replace("Dr.", "").split(",")[0].strip()

                        # print(display_name)
                        address = provider['address']['line1']
                        phone_number = provider['displayOffice']['phoneNumbers'][0]
                        rating = provider['surveyOverallRatingScore']
                        gender = provider['gender']
                        specialty = provider['specialistDesc']
                        zip_code = provider['address']['line2'][-5:]
                        city = provider['address']['line2'].split(",")[0]
                        state = provider['address']['line2'].split(",")[1][1:-6]
                        npi = provider['npi']

                        # Check if 'age' key exists in provider's information
                        if 'age' in provider:
                            age = provider['age']
                        else:
                            age = None  # Assign a default value or handle as needed

                        important_details = {
                            "NPI": npi, "PROVIDER": display_name, "ADDRESS": address, "CITY": city,
                            "STATE": state, "ZIP_CODE": zip_code, "PHONE_NUMBER": phone_number,
                            "RATING": rating, "GENDER": gender, "SPECIALTY": specialty, "AGE": age
                        }

                # df.to_excel("F:\\Python Code\\POC_API\\Healthgrades\\Healthgrades_doctor.xlsx")
                        important_details_non = {
                            "NPI": "", "PROVIDER": "", "ADDRESS": "", "CITY": "",
                            "STATE": "", "ZIP_CODE": "", "PHONE_NUMBER": "",
                            "RATING": "", "GENDER": "", "SPECIALTY": "", "AGE": ""
                        }
                        df_hel = pd.DataFrame(important_details)
                        if df_hel.empty:
                            df_output = df_output._append(important_details_non, ignore_index=True)
                        else:
                            df_output = df_output._append(df_hel, ignore_index=True)


                # Save the output DataFrame to a new Excel file
                #     df_output.drop_duplicates(subset=df.columns[0], keep='first', inplace=True)

                    output_file = "F:\\Python Code\\POC_API\\Healthgrades\\Healthgrades_doctor_details.xlsx"
                    df_output.to_excel(output_file, index=False)
                    # print(filtered_df)
                    # df_output.drop_duplicates(subset=df.columns[0], keep='first', inplace=True)

                    # df_output = filtered_df
                    # print(df_output.to_markdown())
                    df_output['PROVIDER'] = df_output['PROVIDER'].str.upper()
                    df_output['PROVIDER'] = df_output['PROVIDER'].apply(replace_abbreviations)

                    df['PROVIDER'] = df['PROVIDER'].str.upper()
                    df['PROVIDER'] = df['PROVIDER'].apply(replace_abbreviations)
                    df['New'] = df['PROVIDER'].str.strip()
                    # df['New'] = df['New'].str.upper()
                    # df['New'] = df['PROVIDER'].replace('\s+', '', regex=True)
                    df['New'] = df.New.apply(lambda ch: remove_char(ch))

                    df_new = df['New']
                    df_new = df_new.to_frame(name='New')
                    df_new['New'] = df['New'].str.upper()
                    # df_new['index_col'] = df['index_col']
                    df_new["New"] = df_new["New"].replace('\s+', '', regex=True)
                    df_new['New'] = df_new.New.apply(lambda ch: remove_char(ch))

                    final = pd.concat([df, df_new], axis=1)

                    final = final.loc[:, ~final.columns.duplicated()]
                    # print(final.to_markdown())

                    df_output['New'] = df_output['PROVIDER']
                    df_output['New'] = df_output['New'].astype(str)
                    df_output['New'] = df_output['New'].str.upper()
                    # df_output["New"] = df_output["New"].replace('\s+', '', regex=True)
                    df_output["New"] = df_output.New.apply(lambda ch: remove_char(ch))
                    # print(df_output.to_markdown())
                    # Vlookup = pd.merge(df_new, df_output, on=['New'], how='inner')
                    # final_output = Vlookup.drop_duplicates()
                    # print(final_output.to_markdown())
                    df_final = pd.merge(final, df_output, on=['New'], how='left')
                    # print(df_final.to_markdown())
                    df_final = df_final.drop(["New"], axis=1)
                    df_final.rename(columns={'NPI_x': 'NPI', 'STATE_x': 'STATE', 'PROVIDER_x': 'PROVIDER', 'ADDRESS_x': 'ADDRESS',
                                             'CITY_x': 'CITY', 'ZIP_CODE_x': 'ZIP_CODE', 'PHONE_NUMBER_x': 'PHONE_NUMBER'}, inplace=True)
                    # print(df_final.to_markdown())
                    df_ph = df_final.groupby("PROVIDER")["PHONE_NUMBER"].apply(lambda x: x.unique()).reset_index()
                    df_add = df_final.groupby("PROVIDER")["ADDRESS"].apply(lambda x: x.unique()).reset_index()
                    df_Health = pd.merge(df, df_ph, on="PROVIDER", how='left')

                    df_Health = pd.merge(df_Health, df_add, on="PROVIDER", how='left')

                    # df_Health = df_Health.drop(['New', "PROVIDER"], axis=1)

                    df_Health.rename(columns={'PHONE_NUMBER': 'HEALTHGRADE_PHONE_NO'}, inplace=True)

                    df_Health.rename(columns={'ADDRESS': 'HEALTHGRADE_ADDRESS'}, inplace=True)

                    df_Health = df_Health.fillna("")

                    df_Health.to_excel(temp_full_path + '\\Matched_Healthgrade_Output.xlsx', index=False, header=True)
                    # print(df_Health.to_markdown())
                print("Healthgrade Website Module completed!")
                labeld.configure(text="Healthgrade Website Module completed!")
        #
                progress()
        #
                root.update()
    #
            except Exception as e:

                labeld.configure(text="Healthgrade Website Module completed!", fg="brown")
                progress()
                root.update()
# traceback.print_exc(file=sys.stdout)
#
#
        def psychology():
            try:
        #
                print("\n***Psychology Today Module has started***")

                start_time = datetime.now()
        #


                # df_input = pd.read_excel(input_file)
                df_output = pd.DataFrame()
                for index, row in df.iterrows():
                    try:
                        provider_name = row["PROVIDER"]
                        # provider_name = ["Orion Behavioral Health Network"]
                        # try:
                        url = f"https://www.psychologytoday.com/us/therapists?search={provider_name}"

                        querystring = {"lang": "en"}

                        payload = ""
                        headers = {
                            "sec-ch-ua-mobile": "?0",
                            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36",

                        }

                        response = requests.request("GET", url, data=payload, headers=headers, params=querystring)
                        # print(response.text)
                        seen_profiles = set()

                        if response.status_code == 200:
                            soup = BeautifulSoup(response.content, 'html.parser')
                            doctor_listings = soup.find_all('div', class_='results-row')
                            # print(doctor_listings)

                        for doctor_listing in doctor_listings:
                            profile_url = doctor_listing.find('a', class_='profile-title')['href']
                            seen_profiles.add(profile_url)

                            profile_response = requests.get(profile_url, headers=headers)
                            if profile_response.status_code == 200:
                                profile_soup = BeautifulSoup(profile_response.content, 'html.parser')
                                name_element = profile_soup.find('h1', class_='heading-element profile-title heading-element-2')
                                name = name_element.text.strip() if name_element else "N/A"
                                name = name.upper()
                                type_element = profile_soup.find('span', class_='profile-suffix glossary-tooltip-link',
                                                                 attrs={'data-x': 'profile-suffix-profile-type'})
                                type = type_element.text.strip() if type_element else "N/A"
                                phone_number = profile_soup.find('a', class_='lets-connect-phone-number').text.strip()
                                address = profile_soup.find('span', class_='address-region').text.strip()
                                address_element = profile_soup.find('div', class_='address-line')
                                address_one = address_element.find('span', recursive=False).text.strip() if address_element else "N/A"
                                speciality = profile_soup.find('div', class_='link-row')
                                speciality_1 = speciality.text.strip() if speciality else "N/A"
                                fees_element = profile_soup.find('div', class_='at-a-glance_row--fees')
                                fees = fees_element.text.strip() if fees_element else "N/A"
                                expertise_list = profile_soup.find_all('span', class_='attribute_base')
                                expertise = [item.text.strip() for item in expertise_list][:3]

                                # print(f"Name: {name}")
                                # print(f"Type : {type}")
                                # print(f"Phone Number: {phone_number}")
                                # print(f"Zip: {address}")
                                # print(f"Address: {address_one}")
                                # print(f"Speciality: {speciality}")
                                # print(f"Fees : {fees}")
                                # print(f"Expertise: {', '.join(expertise)}")
                        important_details = {
                            "PROVIDER": name, "TYPE": type, "ADDRESS": address_one, "CITY & ZIP": address, "FEES": fees,
                            "EXPERTISE": ', '.join(expertise),
                            "PHONE_NUMBER": phone_number, "SPECIALTY": speciality_1}
                        important_details_non = {
                            "PROVIDER": "", "TYPE": "", "ADDRESS": "", "CITY & ZIP": "", "FEES": "",
                            "EXPERTISE": "",
                            "PHONE_NUMBER": "", "SPECIALTY": ""}
                        # writer.writerow([display_name, address,phone_number, rating,  specialty ])
                        # print(important_details)
                        important_details = {key.upper(): value for key, value in important_details.items()}

                        # print(provider_name)
                        # print(name)
                        if provider_name == name:
                            df_output = df_output._append(important_details, ignore_index=True)
                        else:
                            df_output = df_output._append(important_details_non, ignore_index=True)
                    except:
                        important_details_non = {
                            "PROVIDER": "", "TYPE": "", "ADDRESS": "", "CITY & ZIP": "", "FEES": "",
                            "EXPERTISE": "",
                            "PHONE_NUMBER": "", "SPECIALTY": ""}
                        df_output = df_output._append(important_details_non, ignore_index=True)
                        # print(df_output)
                        # Save the output DataFrame to a new Excel file
                output_file = "F:\\Python Code\\POC_API\\PsychologyToday\\Psychologytoday_doctor_details.xlsx"
                df_output.to_excel(output_file, index=False)
                # print(df_output.to_markdown())


                df_output['PROVIDER'] = df_output['PROVIDER'].str.upper()
                df_output['PROVIDER'] = df_output['PROVIDER'].apply(replace_abbreviations)

                df['PROVIDER'] = df['PROVIDER'].str.upper()
                df['PROVIDER'] = df['PROVIDER'].apply(replace_abbreviations)
                df['New'] = df['PROVIDER'].str.strip()
                # df['New'] = df['New'].str.upper()
                # df['New'] = df['PROVIDER'].replace('\s+', '', regex=True)
                df['New'] = df.New.apply(lambda ch: remove_char(ch))

                df_new = df['New']
                df_new = df_new.to_frame(name='New')
                df_new['New'] = df['New'].str.upper()
                # df_new['index_col'] = df['index_col']
                df_new["New"] = df_new["New"].replace('\s+', '', regex=True)
                df_new['New'] = df_new.New.apply(lambda ch: remove_char(ch))

                final = pd.concat([df, df_new], axis=1)

                final = final.loc[:, ~final.columns.duplicated()]
                # print(final.to_markdown())

                df_output['New'] = df_output['PROVIDER']
                df_output['New'] = df_output['New'].astype(str)
                df_output['New'] = df_output['New'].str.upper()
                # df_output["New"] = df_output["New"].replace('\s+', '', regex=True)
                df_output["New"] = df_output.New.apply(lambda ch: remove_char(ch))
                # print(df_output.to_markdown())
                # Vlookup = pd.merge(df_new, df_output, on=['New'], how='inner')
                # final_output = Vlookup.drop_duplicates()
                # print(final_output.to_markdown())
                df_final = pd.merge(final, df_output, on=['New'], how='left')
                # print(df_final.to_markdown())
                df_final = df_final.drop(["New"], axis=1)
                df_final.rename(columns={'NPI_x': 'NPI', 'STATE_x': 'STATE', 'PROVIDER_x': 'PROVIDER', 'ADDRESS_x': 'ADDRESS',
                                         'CITY_x': 'CITY', 'ZIP_CODE_x': 'ZIP_CODE', 'PHONE_NUMBER_x': 'PHONE_NUMBER'}, inplace=True)
                # print(df_final.to_markdown())
                df_ph = df_final.groupby("PROVIDER")["PHONE_NUMBER"].apply(lambda x: x.unique()).reset_index()
                df_add = df_final.groupby("PROVIDER")["ADDRESS"].apply(lambda x: x.unique()).reset_index()

                df_psy = pd.merge(df, df_ph, on="PROVIDER", how='left')

                df_psy = pd.merge(df_psy, df_add, on="PROVIDER", how='left')

                # df_psy = df_psy.drop(['New', "PROVIDER"], axis=1)

                df_psy.rename(columns={'PHONE_NUMBER': 'PSYCHOLOGYTODAY_PHONE_NO'}, inplace=True)

                df_psy.rename(columns={'ADDRESS': 'PSYCHOLOGYTODAY_ADDRESS'}, inplace=True)

                df_psy = df_psy.fillna("")

                df_psy.to_excel(temp_full_path + '\\Matched_Psychology_Output.xlsx', index=False, header=True)

                # end
                # time = datetime.now()
                print("Psychologytoday Website Module completed successfully!")

                labele.configure(text="Psychologytoday Website Module completed !")

                progress()

                root.update()
            except Exception as e:

                labele.configure(text="Psychologytoday Website Module completed!", fg="brown")

                progress()

                root.update()

                traceback.print_exc(file - sys.stdout)
#
# #
        def datamanipulation():
            try:

                print("\n***Datamanipulation has started***")

                try:

                    NPPES_read = pd.read_excel(temp_full_path + "\\Matched_NPPES_Output.xlsx")

                    NPPES_read.PROVIDER = NPPES_read.PROVIDER.str.upper()

                    NPPES_read.NPPES_ADDRESS = NPPES_read.NPPES_ADDRESS.str.strip('[]').str.strip("''").str.replace(',', '').str.strip('"')

                    NPPES_read.NPPES_ADDRESS = NPPES_read.NPPES_ADDRESS.str.replace("] \[", " , ")
                    NPPES_read.NPPES_ADDRESS = NPPES_read.NPPES_ADDRESS.str.replace("]\[", ";").str.replace('-', '')

                    # for i in range(len(NPPES_read. NPPES Address))2

                    NPPES_read.NPPES_PHONE_NO = (
                        (NPPES_read.NPPES_PHONE_NO.str.strip('[]')).str.strip("''")).str.strip().str.strip('""').str.strip("-")

                    NPPES_read.NPPES_PHONE_NO = NPPES_read.NPPES_PHONE_NO.str.replace('-', "")
                    NPPES_read.NPPES_PHONE_NO = NPPES_read.NPPES_PHONE_NO.str.replace(" ", " ")

                    NPPES_read.NPPES_PHONE_NO = NPPES_read.NPPES_PHONE_NO.str.replace("],\[", " ; ")

                    NPPES_merge = NPPES_read[['PROVIDER', 'NPPES_PHONE_NO', 'NPPES_ADDRESS']]
                    # print(NPPES_merge.to_markdown())

                except:

                    traceback.print_exc(file=sys.stdout)

                    NPPES_merge = pd.DataFrame(columns=['PROVIDER', 'NPPES_PHONE_NO', 'NPPES_ADDRESS'])

                # input = pd.read_excel (input3.get(), sheet_name=current_var_get())

                input = df

                input.PROVIDER = input.PROVIDER.str.upper()
                input = input.astype(str)
                NPPES_merge = NPPES_merge.astype(str)

                join1 = pd.merge(input, NPPES_merge, on=['PROVIDER'], how='left')

                # working with Hippospace

                try:

                    Ind_Read_Hipaa = pd.read_excel(temp_full_path + '\Matched_Hippa_Org_Output.xlsx')

                except:

                    Ind_Read_Hipaa = pd.DataFrame(
                        columns=["PROVIDER", "HIPPASPACE_PHONE_NO", "HIPPASPACE_ADDRESS"])

                final_Hipaa = Ind_Read_Hipaa
                # print(final_Hipaa.to_markdown())

                final_Hipaa.PROVIDER = final_Hipaa.PROVIDER.str.upper()

                final_Hipaa.HIPPASPACE_ADDRESS = (
                    (final_Hipaa.HIPPASPACE_ADDRESS.str.strip('[]').str.strip("''").str.replace(',', '').str.strip('"')))
                final_Hipaa.HIPPASPACE_PHONE_NO = final_Hipaa.HIPPASPACE_PHONE_NO.str.strip('[]').str.strip("''").str.strip().str.strip('""').str.strip("-")
                final_Hipaa.HIPPASPACE_PHONE_NO = final_Hipaa.HIPPASPACE_PHONE_NO.str.replace('\W', '', regex=True)


                final_Hipaa.HIPPASPACE_PHONE_NO = final_Hipaa.HIPPASPACE_PHONE_NO.str.replace("'", "")
                if "'" in final_Hipaa.HIPPASPACE_PHONE_NO:
                    final_Hipaa.HIPPASPACE_PHONE_NO = final_Hipaa.HIPPASPACE_PHONE_NO.str.replace("'", "")


                # print(final_Hipaa.HIPPASPACE_PHONE_NO.to_markdown())
                df_phone_Hipaa = final_Hipaa.groupby("PROVIDER")[("HIPPASPACE_PHONE_NO")].apply(lambda x: x.unique()).reset_index()

                df_Hipaaspace = pd.merge(join1, df_phone_Hipaa, on="PROVIDER", how='left')

                hippa_add = final_Hipaa.groupby("PROVIDER")["HIPPASPACE_ADDRESS"].apply(lambda x: x.unique()).reset_index()

                df_Hipaaspace = pd.merge(df_Hipaaspace, hippa_add, on="PROVIDER", how='left')
                # print(df_Hipaaspace.to_markdown())

                df_Hipaaspace.HIPPASPACE_ADDRESS = df_Hipaaspace.HIPPASPACE_ADDRESS.astype(str).str.lstrip("nan' '").str.lstrip("'").str.rstrip('nan')

                df_Hipaaspace.HIPPASPACE_ADDRESS = df_Hipaaspace.HIPPASPACE_ADDRESS.str.strip('[]').str.strip("''").str.replace(',', '').str.strip('"')

                # print(df_Hipaaspace.HIPPASPACE_ADDRESS.to_markdown())
                for i in range(len(df_Hipaaspace.HIPPASPACE_ADDRESS)):
                    hyphen_idx = df_Hipaaspace.HIPPASPACE_ADDRESS[i].rfind("-")
                    if hyphen_idx != -1:
                        df_Hipaaspace.loc[i, "HIPPASPACE_ADDRESS"] = df_Hipaaspace.HIPPASPACE_ADDRESS[1][:hyphen_idx]

                    else:

                        space_idx = df_Hipaaspace.HIPPASPACE_ADDRESS[i].rfind(' ', -5)
                        df_Hipaaspace.loc[i, 'HIPPASPACE_ADDRESS'] = df_Hipaaspace.HIPPASPACE_ADDRESS[i][:space_idx]

                df_Hipaaspace.HIPPASPACE_PHONE_NO = (df_Hipaaspace.HIPPASPACE_PHONE_NO.astype(str).str.strip('[]')).astype(str).str.lstrip("nan''").str.lstrip("'").str.rstrip("'")

                    # working with WebMD
                # print(df_Hipaaspace.to_markdown())

                try:

                    WebMD_Ind_Read = pd.read_excel(temp_full_path + '\Matched_Webmd_Org_Output.xlsx')

                except:

                    WebMD_Ind_Read = pd.DataFrame(columns=['PROVIDER', 'WEBMD_PHONE_NO', 'WEBMD_ADDRESS'])
                # print(WebMD_Ind_Read.to_markdown())

                final_webmd = WebMD_Ind_Read

                final_webmd.WEBMD_ADDRESS = (
                    (final_webmd.WEBMD_ADDRESS.str.strip('[]').str.strip("''").str.replace(',', '').str.strip('"')))
                final_webmd.WEBMD_PHONE_NO = final_webmd.WEBMD_PHONE_NO.str.strip('[]').str.strip("''").str.strip().str.strip(
                    '""').str.strip("-")
                final_webmd.WEBMD_PHONE_NO = final_webmd.WEBMD_PHONE_NO.str.replace('\W', '', regex=True)

                df_webmd_phn = final_webmd.groupby("PROVIDER")[('WEBMD_PHONE_NO')].apply(lambda x: x.unique()).reset_index()


                df_WebMd = pd.merge(df_Hipaaspace, df_webmd_phn, on="PROVIDER", how='left')
                df_address_webMD = final_webmd.groupby("PROVIDER")[('WEBMD_ADDRESS')].apply(lambda x: x.unique()).reset_index()

                df_WebMd = pd.merge(df_WebMd, df_address_webMD, on="PROVIDER", how='left')


                df_WebMd.WEBMD_ADDRESS = df_WebMd.WEBMD_ADDRESS.astype(str).str.lstrip("nan''").str.lstrip(
                    "'").str.rstrip('nan')

                df_WebMd.WEBMD_ADDRESS = df_WebMd.WEBMD_ADDRESS.str.strip('[]').str.strip("''").str.replace(',', '').str.strip('"').str.lstrip("nan''").str.strip('-')


                df_WebMd.WEBMD_PHONE_NO = (df_WebMd.WEBMD_PHONE_NO.astype(str).str.strip('[]')).astype(
                    str).str.lstrip("nan''").str.lstrip("'").str.rstrip("'")

                # print(df_WebMd.to_markdown())
                # Healthgrade module started

                try:

                    health_read = pd.read_excel(temp_full_path + '\\Matched_Healthgrade_Output.xlsx')
                    # print(health_read.to_markdown())
                    # health_read["PROVIDER"] = health_read["FIRST NAME"] + " " + health_read["LAST NAME"]
                    # print(health_read.to_markdown())
                except:

                    health_read = pd.DataFrame(columns=['PROVIDER', 'HEALTHGRADE_PHONE_NO', 'HEALTHGRADE_ADDRESS'])

                # health_read.PROVIDER = health_read.PROVIDER.str.upper()

                health_read.HEALTHGRADE_ADDRESS = (

                    (health_read.HEALTHGRADE_ADDRESS.str.strip('[]')).str.strip("''").str.replace(',', '').str.replace('-',
                                                                                                                       '').str.extract(
                        '(\D?\w.*)')[0])

                health_read.HEALTHGRADE_PHONE_NO = (health_read.HEALTHGRADE_PHONE_NO.str.strip('[]')).str.strip("''")

                health_read.HEALTHGRADE_PHONE_NO = health_read.HEALTHGRADE_PHONE_NO.str.replace('\W', '', regex=True)

                df_phone_health = health_read.groupby("PROVIDER")[('HEALTHGRADE_PHONE_NO')].apply(lambda x: x.unique()).reset_index()

                df_add_health = health_read.groupby("PROVIDER")["HEALTHGRADE_ADDRESS"].apply(lambda x: x.unique()).reset_index()

                df_health_final = pd.merge(df_phone_health, df_add_health, on="PROVIDER", how='left')
                join4 = pd.merge(df_WebMd, df_health_final, on="PROVIDER", how='left')


                join4.HEALTHGRADE_ADDRESS = (join4.HEALTHGRADE_ADDRESS.astype(str).str.strip('[]')).astype(str).str.strip("''").str.lstrip("nan''")

                join4.HEALTHGRADE_PHONE_NO = (join4.HEALTHGRADE_PHONE_NO.astype(str).str.strip('[]')).astype(
                    str).str.strip("''").str.lstrip("nan''")
                # print(join4.to_markdown())
                try:

                    psy_read = pd.read_excel(temp_full_path + '\\Matched_Psychology_Output.xlsx')

                except:

                    psy_read = pd.DataFrame(columns=['PROVIDER', 'PSYCHOLOGYTODAY_PHONE_NO', 'PSYCHOLOGYTODAY_ADDRESS'])
                psy_read.PROVIDER = psy_read.PROVIDER.str.upper()
                psy_read.PSYCHOLOGYTODAY_ADDRESS = (psy_read.PSYCHOLOGYTODAY_ADDRESS.str.strip('[]')).str.strip("''").str.replace(',',
                                                                                                                        '').str.replace(
                    '-', '')

                psy_read.PSYCHOLOGYTODAY_ADDRESS = (psy_read.PSYCHOLOGYTODAY_ADDRESS.astype(str).str.strip('[]')).astype(
                    str).str.lstrip(

                    "nan''").str.lstrip("'")

                psy_read.PSYCHOLOGYTODAY_PHONE_NO = (

                    psy_read.PSYCHOLOGYTODAY_PHONE_NO.astype(str).str.strip('[]')).astype(str).str.lstrip(

                    "nan").str.lstrip("'")

                psy_read.PSYCHOLOGYTODAY_PHONE_NO = psy_read.PSYCHOLOGYTODAY_PHONE_NO.str.replace('\W', '', regex=True)

                df_phone_phy = psy_read.groupby("PROVIDER")[('PSYCHOLOGYTODAY_PHONE_NO')].apply(
                    lambda x: x.unique()).reset_index()

                df_add_phy = psy_read.groupby("PROVIDER")["PSYCHOLOGYTODAY_ADDRESS"].apply(
                    lambda x: x.unique()).reset_index()

                df_psy_final = pd.merge(df_phone_phy, df_add_phy, on="PROVIDER", how='left')
                join5 = pd.merge(join4, df_psy_final, on="PROVIDER", how='left')
                join5.PSYCHOLOGYTODAY_ADDRESS = (join5.PSYCHOLOGYTODAY_ADDRESS.astype(str).str.strip('[]')).astype(str).str.strip(
                    "''")
                join5.PSYCHOLOGYTODAY_PHONE_NO = (join5.PSYCHOLOGYTODAY_PHONE_NO.astype(str).str.strip('[]')).astype(
                    str).str.strip("''")
                # join5 = join5.drop_duplicates(keep='first')

                # print(join5.to_markdown())

                join5 = join5.replace('nan', '')
                join5.fillna("", inplace=True)
                temp = join5[join5['PROVIDER'].duplicated() & join5['ZIP'].duplicated()]

                temp = temp.groupby('PROVIDER')[("ZIP")].apply(lambda x: x.unique()).reset_index()

                temp_final = pd.merge(join5, temp, on='PROVIDER', how='left')

                # print(temp final.to string())

                # join5['Comments'] = np.where((temp_final['ZIP_x'] == temp_final['ZIP_y']), "Duplicate Provider found in input sheet, Please review.", "")

                join5 = join5.replace('nan', '')

                join5.fillna("", inplace=True)


                join5.drop_duplicates(inplace=True)

                join5.to_excel(output_full_path + '\Provider_Data_Extracted.xlsx'.split('.xlsx')[0] + '.xlsx',
                               index=False, header=True)
                #
                # x2 = win32com.client.Dispatch("Excel.Application")
                #
                # x2.Visible = False
                #
                # x2.Workbooks.Open(Filename=os.path.join(os.getcwd(), "My_Output_macro.xlsm"))
                #
                # x2.Application.Run("My_Output_macro.xlsm!Macro1")
                #
                # x2.Application.quit()
                #
                # del x2

                scrubbed_df = pd.read_excel(os.path.join(os.getcwd(), r"Output File/Provider_Data_Extracted.xlsx"))
                scrubbed_df["SUITE"] = scrubbed_df["SUITE"].replace("", "")

                scrubbed_df["input_address"] = ""

                scrubbed_df["input_address"] = scrubbed_df["input_address"].replace('', '')

                scrubbed_df.fillna("", inplace=True)

                scrubbed_df["ZIP"] = scrubbed_df["ZIP"].astype(str).apply(lambda x: x.zfill(5) if len(x) == 4 else x)

                scrubbed_df["input_address"] = scrubbed_df['STREET'] + " " + scrubbed_df['SUITE'] + " " + scrubbed_df['CITY'] + " " + scrubbed_df['STATE'] + " " + scrubbed_df["ZIP"]

                clean_address = lambda x: re.sub(r'\s+', ' ', x)

                scrubbed_df["input_address"] = scrubbed_df["input_address"].apply(clean_address).str.upper()

                # remove extra whitesapces and coverting into uppercase

                address_df = scrubbed_df.loc[:,
                             ["PROVIDER", "input_address", "NPPES_ADDRESS", "HIPPASPACE_ADDRESS", "WEBMD_ADDRESS",
                              "HEALTHGRADE_ADDRESS", "PSYCHOLOGYTODAY_ADDRESS"]]

                # print(address_df.to_markdown ())

                cols_to_clean = ["input_address", "NPPES_ADDRESS", "HIPPASPACE_ADDRESS", "WEBMD_ADDRESS",
                              "HEALTHGRADE_ADDRESS", "PSYCHOLOGYTODAY_ADDRESS"]

                address_df[cols_to_clean] = address_df[cols_to_clean].applymap(lambda x: x.upper() if isinstance(x, str) else x)

                address_df.replace('nan', '')

                address_df.fillna("", inplace=True)

                # address_df[cols_to_clean] = address_df[cols_to_clean].applymap(replace_abbreviations)

                address_df.to_excel(os.path.join(os.getcwd(), r'Output File/Addresses.xlsx'),index=False)

                path = temp_full_path

                # Confidence score section for phone number

                temp_df = pd.read_excel(os.path.join(os.getcwd(), r"Output File\Provider_Data_Extracted.xlsx"))

                conf_score_dict = {"PROVIDER": [x for x in temp_df.PROVIDER],

                    "NPPES_Score_Ph": [0 for i in range(len(temp_df.index))],

                    "Hippaspace_Score_Ph": [0 for i in range(len(temp_df.index))],

                    "WebMD_Score_Ph": [0 for i in range(len(temp_df.index))],

                    "Healthgrades_Score Ph": [0 for i in range(len(temp_df.index))],

                    "Psychology_Score Ph": [0 for i in range(len(temp_df.index))],

                    "Cumulative_Score_Ph": [0 for i in range(len(temp_df.index))],

                    "Matched_Score_Ph": [0 for i in range(len(temp_df.index))]

                        }

                conf_score_df = pd.DataFrame(conf_score_dict)

                conf_df = temp_df[["PROVIDER", "PHONE", "NPPES_PHONE_NO", "HIPPASPACE_PHONE_NO", "WEBMD_PHONE_NO",
                                   "HEALTHGRADE_PHONE_NO", "PSYCHOLOGYTODAY_PHONE_NO"]]

                conf_df.to_excel(os.path.join(os.getcwd(), r'Output File/Phone.xlsx'), index=False)

                conf_df = conf_df[["PROVIDER", "PHONE", "NPPES_PHONE_NO", "HIPPASPACE_PHONE_NO", "WEBMD_PHONE_NO",
                                   "HEALTHGRADE_PHONE_NO", "PSYCHOLOGYTODAY_PHONE_NO"]].replace('\D', '', regex=True)

                # conf_df.fillna("", inplace=True)

                # print(conf.df)

                for index, val in conf_df.iterrows():
                    row_val = []

                    for i in range(len(val)):
                        if i > 1:

                            if val[i] != "":
                                conf_score_df.iloc[index, i - 1] = 1

                                conf_score_df.iloc[index, i - 1] = 1

                                conf_score_df.iloc[index, -2] = conf_score_df.iloc[index, -2] + 1

                                # if (str(val[1])).strip() == (str(val[i])).strip():

                                if (str(val[i])).strip().find((str(val[1])).strip()) >= 0:
                                    conf_score_df.iloc[index, -1] = conf_score_df.iloc[index, -1] + 1

                conf_score_df['PROVIDER'] = range(1, len(temp_df) + 1)

                cols = conf_score_df.columns.tolist()

                cols = cols[-1:] + cols[:-1]

                conf_score_df = conf_score_df[cols]

                add_df = pd.read_excel(os.path.join(os.getcwd(), r"Output File\Addresses.xlsx"))


                # print(add_df.to_markdown())
                add_conf_score_dict = {

                    "PROVIDER": [x for x in temp_df.PROVIDER],

                    "NPPES_Score_Add": [0 for i in range(len(temp_df.index))],

                    "Hippaspace_Score_Add": [0 for i in range(len(temp_df.index))],

                    "WebMD_Score_Add": [0 for i in range(len(temp_df.index))],

                    "Healthgrades_Score_Add": [0 for i in range(len(temp_df.index))],

                    "Psychology_Score_Add": [0 for i in range(len(temp_df.index))],

                    "Cumulative_Score_Add": [0 for i in range(len(temp_df.index))],

                    "Matched_Score_Add": [0 for i in range(len(temp_df.index))]

                }

                add_conf_score_df = pd.DataFrame(add_conf_score_dict)

                # print((pd.DataFrame(add_conf_score_dict)).to_markdown())

                conf_df = add_df

                conf_df[["PROVIDER", "input_address", "NPPES_ADDRESS", "HIPPASPACE_ADDRESS", "WEBMD_ADDRESS", "HEALTHGRADE_ADDRESS", "PSYCHOLOGYTODAY_ADDRESS"]] = conf_df[["PROVIDER", "input_address", "NPPES_ADDRESS",
                                "HIPPASPACE_ADDRESS", "WEBMD_ADDRESS", "HEALTHGRADE_ADDRESS", "PSYCHOLOGYTODAY_ADDRESS"]].replace('\s+', '', regex=True)

                conf_df.fillna("", inplace=True)

                # print(conf_df)

                for index, val in conf_df.iterrows():
                    row_val = []

                    for i in range(len(val)):
                        if i > 1:

                            if val[i] != "":
                                add_conf_score_df.iloc[index, i - 1] = 1

                                add_conf_score_df.iloc[index, i - 1] = 1

                                add_conf_score_df.iloc[index, -2] = add_conf_score_df.iloc[index, -2] + 1

                    # if (str(val[1])).strip() == (str(val[1]) Istrip():

                                if (str(val[i])).strip().find((str(val[1])).strip()) >= 0:
                                    add_conf_score_df.iloc[index, -1] = add_conf_score_df.iloc[index, -1] + 1

                add_conf_score_df["PROVIDER"] = range(1, len(add_df) + 1)

                # add_ph_conf_score_df = pd.merge(conf_score_df, add_conf_score_df, on='PROVIDER', how='left')

                # add_ph_conf_score_df = add_ph_conf_score_df.drop(["PROVIDER_y"], axis=1)

                add_ph_conf_score_df = pd.merge(conf_score_df, add_conf_score_df, on='PROVIDER', how='left')

                # add_ph_conf_score_df = add_ph_conf_score_df.drop(["PROVIDER_y"], axis=1)

                add_ph_conf_score_df.rename(columns={'PROVIDER_x': 'PROVIDER'}, inplace=True)

                add_ph_conf_score_df["Confidence_Score_Ph"] = ""

                add_ph_conf_score_df["Confidence_Score_Add"] = ""

                add_ph_conf_score_df["Conf_Score_Ph"] = (add_ph_conf_score_df['Matched_Score_Ph'] /

                                                         add_ph_conf_score_df["Cumulative_Score_Ph"]) * 100

                add_ph_conf_score_df["Conf_Score_Ph"] = add_ph_conf_score_df["Conf_Score_Ph"].astype(str)

                add_ph_conf_score_df["Conf_Score_Ph"] = add_ph_conf_score_df["Conf_Score_Ph"] + "%"

                add_ph_conf_score_df["Conf_Score_Add"] = (add_ph_conf_score_df["Matched_Score_Add"] /

                                                             add_ph_conf_score_df["Cumulative_Score_Add"]) * 100

                add_ph_conf_score_df["Conf_Score_Add"] = add_ph_conf_score_df["Conf_Score_Add"].astype(str)

                add_ph_conf_score_df["Conf_Score_Add"] = add_ph_conf_score_df["Conf_Score_Add"] + "%"

                add_ph_conf_score_df["Conf_Score_Ph"] = add_ph_conf_score_df["Conf_Score_Ph"].replace('nan%', '0.0%')

                add_ph_conf_score_df["Conf_Score_Add"] = add_ph_conf_score_df["Conf_Score_Add"].replace('nan%', '0.0%')

                add_ph_conf_score_df.loc[

                    add_ph_conf_score_df['Cumulative_Score_Ph'] == 0, 'Confidence_Score_Ph'] = "unable to validate"

                add_ph_conf_score_df.loc[add_ph_conf_score_df['Cumulative_Score_Ph'] != 0, 'Confidence_Score_Ph'] = add_ph_conf_score_df["Conf_Score_Ph"]

                add_ph_conf_score_df.loc[

                    add_ph_conf_score_df['Cumulative_Score_Add'] == 0, "Confidence_Score_Add"] = "unable to validate"
                add_ph_conf_score_df.loc[add_ph_conf_score_df['Cumulative_Score_Add'] != 0, 'Confidence_Score_Add'] = add_ph_conf_score_df["Conf_Score_Add"]

                conf_df = add_ph_conf_score_df[["PROVIDER", "Confidence_Score_Ph", "Confidence_Score_Add"]]

                # print(conf_df)

                add_ph_conf_score = df.to_excel(os.path.join(os.getcwd(), "Output File\ConfidenceScore.xlsx"), index=False, header=True)

                # conf_df['PROVIDER'] = conf_df['PROVIDER'].astype(str)

                final_df_cons = pd.concat([temp_df, conf_df], axis=1)#, on='PROVIDER', how='left')


                final_df_cons = final_df_cons.loc[:, ~final_df_cons.columns.duplicated()]
                # final_df_cons = final_df_cons.drop_duplicates(subset='NEW')

                final_df_cons = final_df_cons.drop(['New'], axis=1)
                final_df_cons = final_df_cons.fillna("")
                # print(final_df_cons.to_markdown())
                final_df_cons.rename(columns={'Confidence_Score_Ph': 'Phone_Confidence_Score',
                                              'Confidence_Score_Add': 'Address_Confidence_Score'}, inplace=True)


                final_df_cons.to_excel(os.path.join(os.getcwd(), r"Output File\Provider_Data_Extracted_Updated.xlsx"), index=False, header=True)

                input_df = pd.read_excel(output_full_path + '\\Provider_Data_Extracted_Updated_1.xlsx')

                new_join = pd.DataFrame(input_df)
                new_join.fillna("", inplace=True)
                new_join['NPPES_PHONE_NO'] = new_join['NPPES_PHONE_NO'].astype(str).str.replace('.0', '')
                new_join['WEBMD_PHONE_NO'] = new_join['WEBMD_PHONE_NO'].astype(str).str.replace('.0', '')
                new_join['HIPPASPACE_PHONE_NO'] = new_join['HIPPASPACE_PHONE_NO'].astype(str).str.replace('.0', '')
                new_join['HEALTHGRADE_PHONE_NO'] = new_join['HEALTHGRADE_PHONE_NO'].astype(str).str.replace('.0', '')
                new_join['PSYCHOLOGYTODAY_PHONE_NO'] = new_join['PSYCHOLOGYTODAY_PHONE_NO'].astype(str).str.replace('.0', '')
                # new_join = final_df_cons
                # new_join["HIPPASPACE_PHONE_NO"] = new_join["HIPPASPACE_PHONE_NO"].astype(int)


                new_join = new_join.drop(
                    ['NPPES_PHONE_NO', 'NPPES_ADDRESS', 'HIPPASPACE_PHONE_NO', 'HIPPASPACE_ADDRESS', "WEBMD_PHONE_NO",
                    'WEBMD_ADDRESS', 'HEALTHGRADE_PHONE_NO', 'HEALTHGRADE_ADDRESS', 'PSYCHOLOGYTODAY_PHONE_NO',
                                                                    'PSYCHOLOGYTODAY_ADDRESS'], axis=1)
                # new_join = new_join.astype(int)
                # for f in glob.iglob(path + '/**/*.xlsx', recursive=True):
                #     #
                #     os.remove(f)


                clear_treeview()

                tree["column"] = list(new_join.columns)

                tree["show"] = "headings"


                for col in tree["column"]:
                # tree.column (col, stretch=NO)

                    tree.heading(col, text=col)  # Put Data in Rows

                    df_rows = new_join.to_numpy().tolist()

                for row in df_rows:
                    tree.insert("", "end", values=row)


                for item in tree.get_children():
                    phone_score = float(
                        tree.item(item, "values")[-2].strip('%'))  # Assuming Phone_Confidence_Score is the second to last column
                    address_score = float(
                        tree.item(item, "values")[-1].strip('%'))  # Assuming Address_Confidence_Score is the last column

                    # Configure tags for Phone_Confidence_Score
                    if 80 <= phone_score <= 100:
                        phone_tag = "green"
                    elif 60 <= phone_score < 79:
                        phone_tag = "yellow"
                    else:
                        phone_tag = "red"

                    tree.tag_configure(phone_tag, background=phone_tag)

                    # Apply the Phone_Confidence_Score tag to the cell
                    tree.item(item, tags=(phone_tag,))



                    # Configure tags for Address_Confidence_Score
                    if 80 <= address_score <= 100:
                        address_tag = "green"
                    elif 60 <= address_score < 79:
                        address_tag = "yellow"
                    else:
                        address_tag = "red"

                    tree.tag_configure(address_tag, background=address_tag)

                    # Apply the Address_Confidence_Score tag to the cell
                    tree.item(item, tags=(address_tag,))

                # messagebox.showinfo("Information Only", "Processing Completed")

                labelh.configure(text="Final Result Prepared!")
                labelz.configure(text="Current Status--Completed", fg="green")

                progress()

                root.update()

                messagebox.showinfo("Information Only", "Processing Completed")

            except Exception as e:

                labelf.configure(text="Final Result Prepared!", fg="brown")

                labelz.configure(text="Current Status--Completed with error", fg="green")

                progress()

                root.update()

                traceback.print_exc(file=sys.stdout)

                print("We are having some issue while combining Outputs from Different Modules, Please Re-Run the EXE File")
                messagebox.showinfo("Information Only", "Processing Completed with Error")



        def cleanup_temp():

            try:

                source_dir = temp_full_path

                archive_full_path = os.path.join(application_path, archive_file)

                file_names = os.listdir(source_dir)

                for file in file_names:
                    shutil.move(os.path.join(source_dir, file), os.path.join(archive_full_path, file))

            except Exception as e:

                traceback.print_exc(file=sys.stdout)

        def cleanup_out():
            try:

                source_dir = output_full_path

                archive_full_path = os.path.join(application_path, archive_file)

                file_names = os.listdir(source_dir)

                for file in file_names:

                    if file != "Provider_Data_Extracted_Updated_1.xlsx":
                        shutil.move(os.path.join(source_dir, file), os.path.join(archive_full_path, file))

            except Exception as e:

                traceback.print_exc(file - sys.stdout)

        def master_function():
            Nppes()
            Hippaspace()
            Webmd()
            Healthgrades()
            psychology()
            datamanipulation()
            cleanup_temp()
            cleanup_out()
        if __name__ == "__main__":
            Nppes()
            Hippaspace()
            Webmd()
            Healthgrades()
            psychology()
            datamanipulation()
            cleanup_temp()
            cleanup_out()

        # print('\n' + str(filepath) + '_Processed')

        end_time = datetime.now()

        print("\n Total Script Execution time: {}".format(end_time - start_time))

        end_time = datetime.now()

        print("\n######Code run successful!#################")

        print("\n Please open the Excel Files in the Output Folder!!!")

        print("\n Total Script Execution time: {}".format(end_time - start_time))

    #
def detailview():

    try:

        input_df = pd.read_excel(output_full_path + '\\Provider_Data_Extracted_Updated_1.xlsx')


        df1 = pd.DataFrame(input_df)
        df1.fillna("", inplace=True)
        df1['NPPES_PHONE_NO'] = df1['NPPES_PHONE_NO'].astype(str).str.replace('.0', '')
        df1['WEBMD_PHONE_NO'] = df1['WEBMD_PHONE_NO'].astype(str).str.replace('.0', '')
        df1['HIPPASPACE_PHONE_NO'] = df1['HIPPASPACE_PHONE_NO'].astype(str).str.replace('.0', '')
        df1['HEALTHGRADE_PHONE_NO'] = df1['HEALTHGRADE_PHONE_NO'].astype(str).str.replace('.0', '')
        df1['PSYCHOLOGYTODAY_PHONE_NO'] = df1['PSYCHOLOGYTODAY_PHONE_NO'].astype(str).str.replace('.0', '')
        df1 = df1.applymap(lambda x: x.upper() if isinstance(x, str) else x)

        clear_treeview()

        tree["column"] = list(df1.columns)

        tree["show"] = "headings"

        for col in tree["column"]:
            tree.heading(col, text=col)

            df_rows = df1.to_numpy().tolist()

        for row in df_rows:
            tree.insert("", "end", values=row)

        for item in tree.get_children():
            phone_score = float(
                tree.item(item, "values")[-2].strip(
                    '%'))  # Assuming Phone_Confidence_Score is the second to last column
            address_score = float(
                tree.item(item, "values")[-1].strip('%'))  # Assuming Address_Confidence_Score is the last column

            # Configure tags for Phone_Confidence_Score
            if 80 <= phone_score <= 100:
                phone_tag = "green"
            elif 60 <= phone_score < 79:
                phone_tag = "yellow"
            else:
                phone_tag = "red"

            tree.tag_configure(phone_tag, background=phone_tag)

            # Apply the Phone_Confidence_Score tag to the cell
            tree.item(item, tags=(phone_tag,))

            # Configure tags for Address_Confidence_Score
            if 80 <= address_score <= 100:
                address_tag = "green"
            elif 60 <= address_score < 79:
                address_tag = "yellow"
            else:
                address_tag = "red"

            tree.tag_configure(address_tag, background=address_tag)

            # Apply the Address_Confidence_Score tag to the cell
            tree.item(item, tags=(address_tag,))


    except Exception as e:

        traceback.print_exc(file=sys.stdout)

label1 = tk.Label(root, text="Provider Data Search", font=("Cabbria 20 bold"), fg="magenta4", relief=FLAT)

# exitbutton Button (root, text="Exit", width=10, bg="brown", fg="white", command-closeBtn)
from PIL import ImageTk, Image  # Pillow for handling various image formats
label2 = tk.Label(root, text="Enter Excel File Path", width=20, bg="light green", font=("arial", 10, "bold"))
# logo = r'F:\Python Code\POC_API\CDM code Anisha\Provider_Serach_API\exl logo\EXL_Service_logo.png'
image = Image.open(r"F:\Python Code\POC_API\CDM code Anisha\Provider_Serach_API\exl logo\EXL_Service_logo.png")
# logo = ImageTk.PhotoImage(image)
resized_image = image.resize((150, 20))  # Resize to 100x100 pixels
logo = ImageTk.PhotoImage(resized_image)

# Create a label to display the logo image
logo_label = tk.Label(root, image=logo)
# logo_label.pack(padx=100, pady=20)
logo_label.place(x=1200, y=10)

searchButton = tk.Button(root, text="Upload Input", width=15, height=1, bg="brown", fg="white", command=open_file)

input3 = tk.Entry(root, textvariable=xl, width=175, state="disabled", disabledbackground="Light yellow")

label3 = tk.Label(root, text="Excel Sheet", width=20, font=("arial", 10, "bold"), bg="light green")

sheetCombo = ttk.Combobox(root, textvariable=current_var, width=65)

sheetButton = tk.Button(root, text="Pick", width=12, bg="brown", fg="white", command=sheet_data)

# headButton = tk.Button(root, text="Submit", width=15, bg="brown", fg="white", command= renameCol)

Output = tk.Button(root, text="Export Output", width=15, bg="brown", fg="white", command=exportfile)

Submit = tk.Button(root, text="Search Contact Info", width=15, bg="brown", fg="white", command=code_search)

tblFrame = tk.Frame(root, bg="light yellow", bd=2, height=380, width=1350, relief="groove")

pb = ttk.Progressbar(root, orient='horizontal', mode='determinate', length=400, maximum=100, value=0)

pb.place(x=190, y=500)

value_label = Label(root, text = update_progress_label())

value_label.place(x=600, y=500)

labelx = tk.Label(root, width=65, font=("arial", 10))

labelz = tk.Label(root, font=("arial", 10, "bold"))

labela = tk.Label(root, font=("arial", 8))

labelb = tk.Label(root, font=("arial", 8))

labelc = tk.Label(root, font=("arial", 8))

labeld = tk.Label(root, font=("arial", 8))

labele = tk.Label(root, font=("arial", 8))

labelf = tk.Label(root, font=("arial", 8))

labelg = tk.Label(root, font=("arial", 8))

labelh = tk.Label(root, font=("arial", 8))

detailedButton = tk.Button(root, text="Detailed View", width=15, bg="brown", fg="white", command=detailview)

label1.place(x=2, y=7)

# exitbutton.place(x=650, y=8)

label2.place(x=7, y=46)
searchButton.place(x=1240, y=43)

input3.place(x=175, y=46)

label3.place(x=7, y=75)

sheetCombo.place(x=175, y=75)

sheetButton.place(x=600, y=73)

# label4.place(x=448,y=75)

# hd.place(x=539,y=76)

# headButton.place (x-630, y=73)

labelx.place(x=30, y=680)

Output.place(x=1240, y=550)

detailedButton.place(x=1240, y=515)

Submit.place(x=1240, y=73)

labelz.configure(text="Current Status")

labelz.place(x=7, y=500)

labela.place(x=7, y=520)

labelb.place(x=7, y=540)

labelc.place(x=7, y=560)

labeld.place(x=7, y=580)

labele.place(x=7, y=600)

labelf.place(x=7, y=628)

labelg.place(x=7, y=640)

labelh.place(x=7, y=660)

tblFrame.place(x=7, y=100)

tblFrame.pack_propagate(0)

# tblFrame.pack(padx=8, pady=115)

xyz = 1

tree = ttk.Treeview(tblFrame, columns=5, height=18)

tree.bind("<ButtonRelease-1>", selectItem)

root.mainloop()

# except Exception as e:
#     traceback.print_exc(file=sys.stdout)

